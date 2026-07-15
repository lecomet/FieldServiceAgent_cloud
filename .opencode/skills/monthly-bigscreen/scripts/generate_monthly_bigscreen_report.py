#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
import html
import json
import re
import sys
import warnings
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path

try:
    import openpyxl
except ImportError as exc:
    raise SystemExit("缺少依赖 openpyxl，请使用 uv --with openpyxl 运行。") from exc

warnings.filterwarnings("ignore", message="Workbook contains no default style.*", category=UserWarning)

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")


CITY_METRICS = [
    ("价值积分", "value_score", "价值积分"),
    ("评价积分", "eval_score", "评价积分"),
    ("151-装维(量)", "service_151", "151装维量"),
    ("FTTR-H/B-装维(量)", "fttr_hb", "FTTR-H/B装维量"),
    ("人均价值积分", "avg_value_score", "人均价值积分"),
    ("人均评价积分", "avg_eval_score", "人均评价积分"),
    ("人均FTTR-H/B", "avg_fttr_hb", "人均FTTR-H/B"),
]
STAFF_METRICS = [
    ("评价积分", "eval_score", "评价积分"),
    ("原始积分", "raw_score", "原始积分"),
    ("发展积分", "dev_score", "发展积分"),
    ("运营积分", "ops_score", "运营积分"),
    ("全业务量", "total_volume", "全业务量"),
    ("151", "service_151", "151"),
    ("FTTR-H/B", "fttr_hb", "FTTR-H/B"),
]
PRIMARY_CITY_CHARTS = ["value_score", "eval_score", "service_151", "fttr_hb"]
EFFICIENCY_CHARTS = ["avg_value_score", "avg_eval_score", "avg_fttr_hb"]
GLOBAL_STAFF_CHARTS = ["eval_score", "raw_score", "total_volume", "service_151", "fttr_hb"]
TOP_LIMIT = 10


def project_root() -> Path:
    return Path(__file__).resolve().parents[4]


def clean_header(value) -> str:
    return str(value or "").replace("\r", "").replace("\n", "").replace("\t", "").strip()


def normalize_city(value) -> str:
    text = str(value or "").strip()
    return text[:-1] if text.endswith("市") else text


def number(value) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = str(value).strip().replace(",", "").replace("%", "")
    if not text:
        return Decimal("0")
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal("0")


def as_float(value: Decimal) -> float:
    return float(value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def escape(value) -> str:
    return html.escape(str(value or ""), quote=True)


def format_day(day: str) -> str:
    if not day:
        return "最新"
    digits = re.sub(r"\D", "", day)
    if len(digits) != 8:
        return day
    return f"{digits[:4]}-{digits[4:6]}-{digits[6:8]}"


def parse_dates_from_name(path: Path) -> tuple[str, str]:
    match = re.search(r"_(\d{8})_(\d{6})_", path.name)
    if not match:
        return "", ""
    return match.group(1), match.group(2)


def read_excel_rows(path: Path) -> tuple[list[str], list[dict]]:
    if path.suffix.lower() != ".xlsx":
        raise ValueError(f"{path} 文件格式不符合要求：只支持 .xlsx")
    with path.open("rb") as handle:
        signature = handle.read(4)
    if not signature.startswith(b"PK"):
        raise ValueError(f"{path} 不是标准 .xlsx 文件")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        iterator = ws.iter_rows(values_only=True)
        try:
            raw_headers = next(iterator)
        except StopIteration as exc:
            raise ValueError(f"{path} 文件为空：第 1 行必须是字段名") from exc
        headers = [clean_header(value) for value in raw_headers]
        indexes = {header: idx for idx, header in enumerate(headers) if header}
        rows = []
        for raw in iterator:
            if not any(value not in (None, "") for value in raw):
                continue
            rows.append({header: raw[idx] if idx < len(raw) else None for header, idx in indexes.items()})
        return headers, rows
    finally:
        wb.close()


def require_columns(path: Path, headers: list[str], required: list[str]) -> None:
    missing = [name for name in required if name not in headers]
    if missing:
        actual = "、".join(header for header in headers if header) or "无"
        raise ValueError(f"{path} 缺少字段: {', '.join(missing)}；实际字段: {actual}")


def latest_matching(root: Path, patterns: list[str]) -> Path | None:
    candidates = []
    for pattern in patterns:
        candidates.extend(path for path in root.rglob(pattern) if path.is_file() and not path.name.startswith("~$"))
    if not candidates:
        return None
    return max(candidates, key=lambda path: path.stat().st_mtime)


def resolve_input(root: Path, explicit: str | None, patterns: list[str], label: str) -> Path:
    if explicit:
        path = Path(explicit)
        if not path.is_absolute():
            path = root / path
    else:
        path = latest_matching(root / "temp" / "data", patterns)
    if not path or not path.exists():
        raise SystemExit(f"未找到{label} Excel，请使用参数显式指定。")
    return path


def row_metric(row: dict, source: str) -> Decimal:
    return number(row.get(source))


def build_city_data(headers: list[str], rows: list[dict]) -> tuple[list[dict], dict | None]:
    require_columns(Path("地市月累计"), headers, ["地市"])
    metric_sources = {source: (key, label) for source, key, label in CITY_METRICS if source in headers}
    cities = []
    province = None
    for row in rows:
        city = normalize_city(row.get("地市"))
        if not city:
            continue
        item = {"city": city}
        for source, (key, _label) in metric_sources.items():
            item[key] = as_float(row_metric(row, source))
        if city == "全省":
            province = item
        else:
            cities.append(item)
    if not cities:
        raise ValueError("地市月累计文件没有具体地市数据")
    return sorted(cities, key=lambda item: item["city"]), province


def build_staff_data(headers: list[str], rows: list[dict]) -> list[dict]:
    require_columns(Path("正式人员月累计"), headers, ["地市", "区县", "装维姓名", "工号"])
    metric_sources = {source: (key, label) for source, key, label in STAFF_METRICS if source in headers}
    staff = []
    for row in rows:
        city = normalize_city(row.get("地市"))
        name = str(row.get("装维姓名") or "").strip()
        staff_id = str(row.get("工号") or "").strip()
        if not city or not name:
            continue
        item = {
            "city": city,
            "county": str(row.get("区县") or "").strip(),
            "name": name,
            "staff_id": staff_id,
        }
        for source, (key, _label) in metric_sources.items():
            item[key] = as_float(row_metric(row, source))
        staff.append(item)
    if not staff:
        raise ValueError("正式人员月累计文件没有有效人员数据")
    return staff


def top_items(items: list[dict], metric: str, limit: int = TOP_LIMIT) -> list[dict]:
    return sorted(items, key=lambda item: number(item.get(metric)), reverse=True)[:limit]


def sum_metric(items: list[dict], metric: str) -> float:
    return as_float(sum((number(item.get(metric)) for item in items), Decimal("0")))


def build_payload(city_rows: list[dict], province: dict | None, staff_rows: list[dict]) -> dict:
    city_metric_labels = {key: label for _source, key, label in CITY_METRICS}
    staff_metric_labels = {key: label for _source, key, label in STAFF_METRICS}
    staff_by_city: dict[str, list[dict]] = {}
    for row in staff_rows:
        staff_by_city.setdefault(row["city"], []).append(row)

    city_rankings = {
        metric: top_items([row for row in city_rows if metric in row], metric, len(city_rows))
        for metric in [*PRIMARY_CITY_CHARTS, *EFFICIENCY_CHARTS]
    }
    global_staff_rankings = {
        metric: top_items([row for row in staff_rows if metric in row], metric)
        for metric in GLOBAL_STAFF_CHARTS
    }
    city_staff_rankings = {
        city: {
            metric: top_items([row for row in rows if metric in row], metric)
            for metric in GLOBAL_STAFF_CHARTS
        }
        for city, rows in staff_by_city.items()
    }

    province_data = province or {}
    city_count = len(city_rows)
    staff_count = len(staff_rows)
    kpis = [
        {"label": "地市数", "value": city_count, "unit": "个"},
        {"label": "有效人员", "value": staff_count, "unit": "人"},
        {"label": "全省评价积分", "value": province_data.get("eval_score", sum_metric(staff_rows, "eval_score")), "unit": "分"},
        {"label": "全省价值积分", "value": province_data.get("value_score", sum_metric(city_rows, "value_score")), "unit": "分"},
        {"label": "151装维量", "value": province_data.get("service_151", sum_metric(city_rows, "service_151")), "unit": "单"},
        {"label": "FTTR-H/B装维量", "value": province_data.get("fttr_hb", sum_metric(city_rows, "fttr_hb")), "unit": "单"},
    ]

    return {
        "kpis": kpis,
        "cities": city_rows,
        "cityMetricLabels": city_metric_labels,
        "staffMetricLabels": staff_metric_labels,
        "cityRankings": city_rankings,
        "globalStaffRankings": global_staff_rankings,
        "cityStaffRankings": city_staff_rankings,
        "cityStaffCounts": {city: len(rows) for city, rows in staff_by_city.items()},
        "primaryCityCharts": PRIMARY_CITY_CHARTS,
        "efficiencyCharts": EFFICIENCY_CHARTS,
        "globalStaffCharts": GLOBAL_STAFF_CHARTS,
    }


def default_output_path(root: Path, acct_day: str) -> Path:
    suffix = re.sub(r"\D", "", acct_day) if acct_day else datetime.now().strftime("%Y%m%d")
    if len(suffix) != 8:
        suffix = datetime.now().strftime("%Y%m%d")
    return root / "output" / f"各地市随销月累计大屏_{suffix}.html"


def render_html(payload: dict, city_file: Path, staff_file: Path, acct_day: str, month_id: str) -> str:
    data = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    title_date = format_day(acct_day)
    month_text = month_id or (re.sub(r"\D", "", acct_day)[:6] if acct_day else "-")
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>各地市随销月累计大屏_{escape(acct_day or "latest")}</title>
  <style>
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      background: #08111f;
      color: #e8f1ff;
      font-family: "Microsoft YaHei", "PingFang SC", Arial, sans-serif;
    }}
    .screen {{
      width: 1920px;
      min-height: 1080px;
      margin: 0 auto;
      padding: 26px 30px 30px;
      background:
        radial-gradient(circle at 18% 12%, rgba(32, 128, 196, 0.28), transparent 30%),
        radial-gradient(circle at 82% 18%, rgba(17, 185, 129, 0.16), transparent 28%),
        linear-gradient(135deg, #07101d 0%, #0c1d33 52%, #07111f 100%);
    }}
    .topbar {{
      display: grid;
      grid-template-columns: 1fr auto;
      align-items: end;
      gap: 24px;
      padding-bottom: 18px;
      border-bottom: 1px solid rgba(125, 211, 252, 0.38);
    }}
    h1 {{
      margin: 0;
      font-size: 38px;
      line-height: 1.15;
      letter-spacing: 0;
      font-weight: 800;
    }}
    .subtitle {{
      margin-top: 10px;
      color: #9fc7ee;
      font-size: 17px;
    }}
    .meta {{
      text-align: right;
      color: #b9d8f4;
      font-size: 14px;
      line-height: 1.8;
      white-space: nowrap;
    }}
    .kpis {{
      display: grid;
      grid-template-columns: repeat(6, 1fr);
      gap: 12px;
      margin: 18px 0;
    }}
    .kpi {{
      min-height: 104px;
      padding: 16px;
      border: 1px solid rgba(125, 211, 252, 0.24);
      border-radius: 8px;
      background: linear-gradient(180deg, rgba(16, 42, 67, 0.92), rgba(8, 25, 44, 0.88));
      box-shadow: inset 0 1px 0 rgba(255,255,255,0.08);
    }}
    .kpi span {{
      display: block;
      color: #91bfe8;
      font-size: 14px;
      margin-bottom: 10px;
    }}
    .kpi strong {{
      color: #ffffff;
      font-size: 30px;
      line-height: 1.05;
      font-variant-numeric: tabular-nums;
    }}
    .kpi em {{
      color: #7dd3fc;
      font-size: 15px;
      font-style: normal;
      margin-left: 5px;
    }}
    .layout {{
      display: grid;
      grid-template-columns: 1.18fr 0.82fr;
      gap: 16px;
    }}
    .panel {{
      border: 1px solid rgba(125, 211, 252, 0.22);
      border-radius: 8px;
      background: rgba(8, 23, 41, 0.82);
      box-shadow: 0 18px 40px rgba(0,0,0,0.24);
      overflow: hidden;
    }}
    .panel-head {{
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 16px;
      padding: 14px 16px;
      border-bottom: 1px solid rgba(125, 211, 252, 0.18);
      background: rgba(15, 43, 72, 0.78);
    }}
    .panel-head h2 {{
      margin: 0;
      font-size: 19px;
      font-weight: 800;
      letter-spacing: 0;
    }}
    .tabs {{
      display: flex;
      gap: 8px;
      flex-wrap: wrap;
      justify-content: flex-end;
    }}
    button {{
      border: 1px solid rgba(125, 211, 252, 0.36);
      border-radius: 999px;
      color: #cfe8ff;
      background: rgba(12, 32, 55, 0.92);
      padding: 7px 11px;
      font-size: 13px;
      cursor: pointer;
    }}
    button.active {{
      color: #06121f;
      background: #7dd3fc;
      border-color: #7dd3fc;
      font-weight: 800;
    }}
    .chart {{
      padding: 12px 16px 16px;
    }}
    .bar-row {{
      display: grid;
      grid-template-columns: 74px 1fr 92px;
      align-items: center;
      gap: 10px;
      min-height: 32px;
      margin: 7px 0;
      font-size: 14px;
    }}
    .bar-label {{
      color: #dbeafe;
      font-weight: 800;
      white-space: nowrap;
      overflow: hidden;
      text-overflow: ellipsis;
    }}
    .bar-track {{
      height: 18px;
      border-radius: 999px;
      background: rgba(148, 163, 184, 0.20);
      overflow: hidden;
      border: 1px solid rgba(148, 163, 184, 0.12);
    }}
    .bar-fill {{
      height: 100%;
      min-width: 2px;
      border-radius: 999px;
      background: linear-gradient(90deg, #38bdf8, #22c55e);
      box-shadow: 0 0 16px rgba(56, 189, 248, 0.38);
    }}
    .bar-value {{
      text-align: right;
      color: #e0f2fe;
      font-weight: 800;
      font-variant-numeric: tabular-nums;
    }}
    .staff-grid {{
      display: grid;
      grid-template-columns: 1fr 1fr;
      gap: 16px;
      margin-top: 16px;
    }}
    .rank-list {{
      padding: 10px 14px 14px;
    }}
    .rank-row {{
      display: grid;
      grid-template-columns: 34px 74px 1fr 92px;
      align-items: center;
      gap: 8px;
      min-height: 37px;
      border-bottom: 1px solid rgba(125, 211, 252, 0.10);
      font-size: 13px;
    }}
    .rank-no {{
      width: 24px;
      height: 24px;
      line-height: 24px;
      border-radius: 50%;
      text-align: center;
      background: rgba(125, 211, 252, 0.18);
      color: #e0f2fe;
      font-weight: 800;
    }}
    .rank-city {{
      color: #93c5fd;
      font-weight: 800;
      white-space: nowrap;
      overflow: hidden;
      text-overflow: ellipsis;
    }}
    .rank-name {{
      color: #f8fafc;
      white-space: nowrap;
      overflow: hidden;
      text-overflow: ellipsis;
    }}
    .rank-value {{
      text-align: right;
      color: #bbf7d0;
      font-weight: 900;
      font-variant-numeric: tabular-nums;
    }}
    .city-strip {{
      display: grid;
      grid-template-columns: repeat(8, 1fr);
      gap: 8px;
      margin-top: 16px;
    }}
    .city-btn {{
      border-radius: 6px;
      padding: 10px 8px;
      text-align: center;
      color: #dbeafe;
      background: rgba(15, 43, 72, 0.80);
      border: 1px solid rgba(125, 211, 252, 0.18);
    }}
    .city-btn.active {{
      color: #07111f;
      background: linear-gradient(90deg, #7dd3fc, #bbf7d0);
      border-color: transparent;
    }}
    .small-note {{
      color: #8fb9dc;
      font-size: 12px;
      line-height: 1.5;
    }}
  </style>
</head>
<body>
  <main class="screen">
    <header class="topbar">
      <div>
        <h1>各地市随销月累计大屏</h1>
        <div class="subtitle">截至日期：{escape(title_date)}　账期：{escape(month_text)}</div>
      </div>
      <div class="meta">
        <div>生成时间：{escape(generated_at)}</div>
        <div>地市数据：{escape(city_file.name)}</div>
        <div>人员数据：{escape(staff_file.name)}</div>
      </div>
    </header>
    <section class="kpis" id="kpis"></section>
    <section class="layout">
      <div>
        <section class="panel">
          <div class="panel-head">
            <h2>地市累计排名</h2>
            <div class="tabs" id="city-tabs"></div>
          </div>
          <div class="chart" id="city-chart"></div>
        </section>
        <section class="panel" style="margin-top:16px;">
          <div class="panel-head">
            <h2>人均效率对比</h2>
            <div class="tabs" id="efficiency-tabs"></div>
          </div>
          <div class="chart" id="efficiency-chart"></div>
        </section>
      </div>
      <div>
        <section class="panel">
          <div class="panel-head">
            <h2 id="staff-title">个人贡献榜</h2>
            <div class="tabs" id="staff-tabs"></div>
          </div>
          <div class="rank-list" id="staff-rank"></div>
        </section>
        <section class="city-strip" id="city-strip"></section>
        <div class="small-note" style="margin-top:12px;">点击地市可查看该地市个人 TOP10；再次点击“全省”返回全局榜单。</div>
      </div>
    </section>
  </main>
  <script>
    const DATA = {data};
    const fmt = new Intl.NumberFormat('zh-CN', {{ maximumFractionDigits: 2 }});
    let cityMetric = DATA.primaryCityCharts[0];
    let efficiencyMetric = DATA.efficiencyCharts[0];
    let staffMetric = DATA.globalStaffCharts[0];
    let selectedCity = '全省';

    function metricLabel(map, key) {{
      return map[key] || key;
    }}
    function maxValue(rows, metric) {{
      return Math.max(1, ...rows.map(row => Number(row[metric] || 0)));
    }}
    function renderTabs(id, metrics, labels, current, onClick) {{
      const el = document.getElementById(id);
      el.innerHTML = metrics.map(metric => `<button type="button" class="${{metric === current ? 'active' : ''}}" data-metric="${{metric}}">${{metricLabel(labels, metric)}}</button>`).join('');
      el.querySelectorAll('button').forEach(btn => btn.addEventListener('click', () => onClick(btn.dataset.metric)));
    }}
    function renderKpis() {{
      document.getElementById('kpis').innerHTML = DATA.kpis.map(item => `
        <div class="kpi">
          <span>${{item.label}}</span>
          <strong>${{fmt.format(Number(item.value || 0))}}</strong><em>${{item.unit}}</em>
        </div>
      `).join('');
    }}
    function renderBars(id, rows, metric, labelMap) {{
      const max = maxValue(rows, metric);
      document.getElementById(id).innerHTML = rows.map(row => {{
        const value = Number(row[metric] || 0);
        const width = Math.max(2, value / max * 100);
        return `
          <div class="bar-row">
            <div class="bar-label" title="${{row.city}}">${{row.city}}</div>
            <div class="bar-track"><div class="bar-fill" style="width:${{width}}%"></div></div>
            <div class="bar-value">${{fmt.format(value)}}</div>
          </div>
        `;
      }}).join('');
    }}
    function renderStaff() {{
      const rankings = selectedCity === '全省' ? DATA.globalStaffRankings : (DATA.cityStaffRankings[selectedCity] || DATA.globalStaffRankings);
      const rows = rankings[staffMetric] || [];
      document.getElementById('staff-title').textContent = selectedCity === '全省' ? '个人贡献榜' : `${{selectedCity}}个人贡献榜`;
      document.getElementById('staff-rank').innerHTML = rows.map((row, idx) => `
        <div class="rank-row">
          <div class="rank-no">${{idx + 1}}</div>
          <div class="rank-city" title="${{row.city}} / ${{row.county}}">${{row.city}}</div>
          <div class="rank-name" title="${{row.county}} ${{row.name}} ${{row.staff_id}}">${{row.name}}<span style="color:#8fb9dc;"> · ${{row.county}}</span></div>
          <div class="rank-value">${{fmt.format(Number(row[staffMetric] || 0))}}</div>
        </div>
      `).join('');
    }}
    function renderCityStrip() {{
      const cities = ['全省', ...DATA.cities.map(row => row.city)];
      document.getElementById('city-strip').innerHTML = cities.map(city => `
        <button type="button" class="city-btn ${{city === selectedCity ? 'active' : ''}}" data-city="${{city}}">
          ${{city}}${{city === '全省' ? '' : `<br><span style="font-size:11px;color:inherit;opacity:.72">${{DATA.cityStaffCounts[city] || 0}}人</span>`}}
        </button>
      `).join('');
      document.querySelectorAll('.city-btn').forEach(btn => btn.addEventListener('click', () => {{
        selectedCity = btn.dataset.city;
        renderCityStrip();
        renderStaff();
      }}));
    }}
    function setCityMetric(metric) {{
      cityMetric = metric;
      renderTabs('city-tabs', DATA.primaryCityCharts, DATA.cityMetricLabels, cityMetric, setCityMetric);
      renderBars('city-chart', DATA.cityRankings[cityMetric], cityMetric, DATA.cityMetricLabels);
    }}
    function setEfficiencyMetric(metric) {{
      efficiencyMetric = metric;
      renderTabs('efficiency-tabs', DATA.efficiencyCharts, DATA.cityMetricLabels, efficiencyMetric, setEfficiencyMetric);
      renderBars('efficiency-chart', DATA.cityRankings[efficiencyMetric], efficiencyMetric, DATA.cityMetricLabels);
    }}
    function setStaffMetric(metric) {{
      staffMetric = metric;
      renderTabs('staff-tabs', DATA.globalStaffCharts, DATA.staffMetricLabels, staffMetric, setStaffMetric);
      renderStaff();
    }}
    function renderAll() {{
      renderKpis();
      renderTabs('city-tabs', DATA.primaryCityCharts, DATA.cityMetricLabels, cityMetric, setCityMetric);
      renderTabs('efficiency-tabs', DATA.efficiencyCharts, DATA.cityMetricLabels, efficiencyMetric, setEfficiencyMetric);
      renderTabs('staff-tabs', DATA.globalStaffCharts, DATA.staffMetricLabels, staffMetric, setStaffMetric);
      renderBars('city-chart', DATA.cityRankings[cityMetric], cityMetric, DATA.cityMetricLabels);
      renderBars('efficiency-chart', DATA.cityRankings[efficiencyMetric], efficiencyMetric, DATA.cityMetricLabels);
      renderCityStrip();
      renderStaff();
    }}
    renderAll();
  </script>
</body>
</html>
"""


def main() -> None:
    parser = argparse.ArgumentParser(description="生成各地市随销月累计大屏 HTML")
    parser.add_argument("--city-file", help="地市月累计 Excel；默认从 temp/data 自动查找")
    parser.add_argument("--staff-file", help="正式人员月累计 Excel；默认从 temp/data 自动查找")
    parser.add_argument("--acct-day", help="截至日期 YYYYMMDD；未传时优先从地市文件名解析")
    parser.add_argument("--month-id", help="账期 YYYYMM；未传时优先从地市文件名或 acct_day 推断")
    parser.add_argument("--output", help="HTML 输出路径；默认 output/各地市随销月累计大屏_{acct_day}.html")
    args = parser.parse_args()

    root = project_root()
    city_file = resolve_input(
        root,
        args.city_file,
        ["*各地市装维月累计*.xlsx", "*装维月累计_地市汇总*.xlsx"],
        "地市月累计",
    )
    staff_file = resolve_input(
        root,
        args.staff_file,
        ["*正式人员装维月累计*.xlsx"],
        "正式人员月累计",
    )

    city_headers, city_raw_rows = read_excel_rows(city_file)
    staff_headers, staff_raw_rows = read_excel_rows(staff_file)
    city_rows, province = build_city_data(city_headers, city_raw_rows)
    staff_rows = build_staff_data(staff_headers, staff_raw_rows)
    payload = build_payload(city_rows, province, staff_rows)

    parsed_day, parsed_month = parse_dates_from_name(city_file)
    acct_day = args.acct_day or parsed_day
    month_id = args.month_id or parsed_month or (acct_day[:6] if acct_day and len(acct_day) >= 6 else "")

    if args.output:
        output = Path(args.output)
        if not output.is_absolute():
            output = root / output
    else:
        output = default_output_path(root, acct_day)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(render_html(payload, city_file, staff_file, acct_day, month_id), encoding="utf-8")

    print("=== 各地市随销月累计大屏 ===")
    print(f"地市文件: {city_file}")
    print(f"人员文件: {staff_file}")
    print(f"输出文件: {output}")
    print(f"地市行数: {len(city_rows)}")
    print(f"人员行数: {len(staff_rows)}")


if __name__ == "__main__":
    main()
