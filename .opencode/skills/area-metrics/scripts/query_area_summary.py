#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
import sys
import warnings
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

try:
    import openpyxl
except ImportError as exc:
    raise SystemExit("缺少依赖 openpyxl，请使用 uv --with openpyxl 运行。") from exc

warnings.filterwarnings("ignore", message="Workbook contains no default style.*", category=UserWarning)

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")


REQUIRED_COLUMNS = {"地市"}
PREVIEW_GROUPS = [
    ["地市", "终端宽带-装维(量)", "移动-装维(量)", "151-装维(量)", "天翼智屏-装维(量)"],
    ["地市", "FTTR-装维(量)", "FTTR-B-装维(量)", "FTTR-H/B-装维(量)"],
    ["地市", "发展积分", "运营积分", "价值积分", "人均价值积分", "人均FTTR-H/B"],
]
DEFAULT_MAX_OUTPUT_ROWS = 20


def project_root() -> Path:
    return Path(__file__).resolve().parents[4]


def latest_xlsx(directory: Path) -> Path | None:
    files = [path for path in directory.glob("*.xlsx") if path.is_file() and not path.name.startswith("~$")]
    if not files:
        return None
    return max(files, key=lambda path: path.stat().st_mtime)


def clean_header(value) -> str:
    return str(value or "").replace("\r", "").replace("\n", "").replace("\t", "").strip()


def format_fields(headers: list[str]) -> str:
    fields = [header for header in headers if header]
    return "、".join(fields) if fields else "无"


def validate_xlsx_file(path: Path) -> None:
    if path.suffix.lower() != ".xlsx":
        raise ValueError(f"{path} 文件格式不符合要求：只支持 .xlsx")
    with path.open("rb") as handle:
        signature = handle.read(4)
    if not signature.startswith(b"PK"):
        raise ValueError(f"{path} 不是标准 .xlsx 文件。请确认 BI 下载格式为 xlsx，不要把 CSV/HTML/XLS 改名为 .xlsx。")


def read_headers(path: Path, iterator) -> list[str]:
    try:
        raw_headers = next(iterator)
    except StopIteration as exc:
        raise ValueError(f"{path} 文件为空：第 1 行必须是字段名") from exc
    headers = [clean_header(value) for value in raw_headers]
    non_empty = [header for header in headers if header]
    duplicates = sorted({header for header in non_empty if non_empty.count(header) > 1})
    if duplicates:
        raise ValueError(f"{path} 第 1 行字段名重复: {', '.join(duplicates)}\n实际字段: {format_fields(headers)}")
    return headers


def require_columns(path: Path, indexes: dict[str, int], required: set[str], headers: list[str]) -> None:
    missing = sorted(required - set(indexes))
    if missing:
        raise ValueError(f"{path} 缺少字段: {', '.join(missing)}\n实际字段: {format_fields(headers)}")


def normalize_city(value) -> str:
    text = str(value or "").strip()
    return text[:-1] if text.endswith("市") else text


def number(value) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    text = str(value).strip().replace(",", "")
    if not text:
        return Decimal("0")
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal("0")


def text_match(value, query: str) -> bool:
    value_text = normalize_city(value)
    query_text = normalize_city(query)
    return bool(value_text and query_text and (value_text == query_text or query_text in value_text or value_text in query_text))


def load_rows(path: Path) -> tuple[list[str], list[dict]]:
    validate_xlsx_file(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    iterator = ws.iter_rows(values_only=True)
    try:
        headers = read_headers(path, iterator)
        indexes = {header: idx for idx, header in enumerate(headers) if header}
        require_columns(path, indexes, REQUIRED_COLUMNS, headers)

        rows = []
        for raw in iterator:
            row = {header: raw[idx] if idx < len(raw) else None for header, idx in indexes.items()}
            row["地市"] = normalize_city(row.get("地市"))
            rows.append(row)
    finally:
        wb.close()
    return headers, rows


def ensure_area_detail(path: Path, rows: list[dict]) -> None:
    city_names = sorted({normalize_city(row.get("地市")) for row in rows if normalize_city(row.get("地市"))})
    detail_names = [city for city in city_names if city != "全省"]
    if not detail_names:
        actual = "、".join(city_names) if city_names else "无"
        raise ValueError(f"{path} 只读取到全省汇总，没有各地市明细。请检查 BI 下载是否保留地市维度。实际地市值: {actual}")


def ratio_pairs(headers: list[str]) -> dict[str, tuple[str, str]]:
    pairs = {}
    for header in headers:
        if not header.endswith("-装维占比"):
            continue
        product = header[: -len("-装维占比")]
        numerator = f"{product}-装维(量)"
        denominator = f"{product}-全渠道(量)"
        if numerator in headers and denominator in headers:
            pairs[header] = (numerator, denominator)
    return pairs


def aggregate_rows(headers: list[str], rows: list[dict]) -> list[dict]:
    grouped: dict[str, dict] = {}
    ratios = ratio_pairs(headers)
    for row in rows:
        city = normalize_city(row.get("地市"))
        if not city:
            continue
        target = grouped.setdefault(city, {header: None for header in headers})
        target["地市"] = city
        if "地市编码" in headers and not target.get("地市编码"):
            target["地市编码"] = row.get("地市编码")
        if "编码" in headers and not target.get("编码"):
            target["编码"] = row.get("编码")
        for header in headers:
            if header in {"地市", "地市编码", "编码"} or header in ratios:
                continue
            target[header] = number(target.get(header)) + number(row.get(header))

    for row in grouped.values():
        for ratio_col, (num_col, den_col) in ratios.items():
            denominator = number(row.get(den_col))
            row[ratio_col] = number(row.get(num_col)) / denominator if denominator else Decimal("0")
    return list(grouped.values())


def format_value(value) -> str:
    if isinstance(value, Decimal):
        return str(round(float(value), 4)).rstrip("0").rstrip(".")
    if isinstance(value, float):
        return str(round(value, 4)).rstrip("0").rstrip(".")
    return str(value or "")


def markdown_table(headers: list[str], rows: list[dict]) -> str:
    lines = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join("---" for _ in headers) + " |",
    ]
    for row in rows:
        lines.append("| " + " | ".join(format_value(row.get(header)) for header in headers) + " |")
    return "\n".join(lines)


def write_excel(output_path: Path, headers: list[str], rows: list[dict], source_file: Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "地市汇总"
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header) for header in headers])

    meta = wb.create_sheet("来源")
    meta.append(["项目", "值"])
    meta.append(["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    meta.append(["来源文件", str(source_file)])
    meta.append(["地市名称处理", "去掉末尾“市”，同名地市合并；占比按合并后的分子/分母重算"])
    meta.append(["行数", len(rows)])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(output_path)
        return output_path
    except PermissionError:
        fallback = output_path.with_name(f"{output_path.stem}_{datetime.now().strftime('%Y%m%d%H%M%S')}{output_path.suffix}")
        wb.save(fallback)
        return fallback


def main() -> None:
    parser = argparse.ArgumentParser(description="装维地市汇总查询和地市名称标准化")
    parser.add_argument("--city", help="可选，限定地市")
    parser.add_argument("--area-dir", default="temp/data/area")
    parser.add_argument("--output", help="可选，输出 Excel 路径")
    parser.add_argument("--max-output", type=int, default=DEFAULT_MAX_OUTPUT_ROWS)
    args = parser.parse_args()

    root = project_root()
    source_file = latest_xlsx(root / args.area_dir)
    if not source_file:
        raise SystemExit(f"未找到地市汇总 Excel: {root / args.area_dir}")

    headers, raw_rows = load_rows(source_file)
    ensure_area_detail(source_file, raw_rows)
    rows = aggregate_rows(headers, raw_rows)
    if args.city:
        rows = [row for row in rows if text_match(row.get("地市"), args.city)]
    if not rows:
        raise SystemExit("未找到匹配的地市汇总数据。")

    rows.sort(key=lambda row: normalize_city(row.get("地市")))
    saved = None
    if args.output:
        output = Path(args.output)
        if not output.is_absolute():
            output = root / output
        saved = write_excel(output, headers, rows, source_file)

    print("=== 装维地市汇总 ===")
    print(f"来源文件: {source_file}")
    if saved:
        print(f"输出文件: {saved}")
    else:
        print(f"输出文件: 未另存；地市汇总下载文件保留在 {args.area_dir}/。")
    print(f"地市行数: {len(rows)}")
    print("地市名称处理: 去掉末尾“市”，同名地市合并；占比按合并后的分子/分母重算")

    if len(rows) > args.max_output:
        if saved:
            print(f"行数超过 {args.max_output}，不在消息栏罗列；请打开 Excel 查看完整结果。")
        else:
            print(f"行数超过 {args.max_output}，不在消息栏罗列；如需保存格式化副本，请显式传 --output。")
        return

    for group in PREVIEW_GROUPS:
        available = [header for header in group if header in headers]
        if len(available) >= 2:
            print()
            print(markdown_table(available, rows))


if __name__ == "__main__":
    main()
