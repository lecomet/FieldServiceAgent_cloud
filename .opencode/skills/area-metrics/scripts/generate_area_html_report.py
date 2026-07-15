#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
import html
import re
import sys
import warnings
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path

try:
    import openpyxl
except ImportError as exc:
    raise SystemExit("缺少依赖 openpyxl，请使用 uv --with openpyxl 运行。") from exc

warnings.filterwarnings("ignore", message="Workbook contains no default style.*", category=UserWarning)

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")


GROUPS = [
    ("移动", ["全渠道(量)", "装维(量)", "装维占比"]),
    ("终端宽带", ["全渠道(量)", "装维(量)", "装维占比"]),
    ("FTTR", ["全渠道(量)", "装维(量)", "装维占比"]),
    ("FTTR-B", ["全渠道(量)", "装维(量)", "装维占比"]),
    ("FTTR-H/B", ["全渠道(量)", "装维(量)", "装维占比"]),
    ("151", ["全渠道(量)", "装维(量)", "装维占比"]),
    ("天翼智屏", ["全渠道(量)", "装维(量)", "装维占比"]),
]
SINGLE_COLUMNS = [
    "加载积分",
    "发展积分",
    "运营积分",
    "价值积分",
    "评价积分",
    "总人数含班长",
    "总人数不含班长",
    "人均价值积分",
    "人均评价积分",
    "人均FTTR-H/B",
]
HIGHLIGHT_COUNT = 3
FOCUS_LIMIT = 5
FOCUS_METRICS = [
    {"key": "151-装维(量)", "title": "151装维量", "kind": "number"},
    {"key": "FTTR-H/B-装维(量)", "title": "FTTR-H/B装维量", "kind": "number"},
    {"key": "人均价值积分", "title": "人均价值积分", "kind": "number"},
]
REQUIRED_COLUMNS = {"地市"}
REPORT_KINDS = {
    "standard": {
        "title": "地市随销统计报表",
        "html_title": "地市随销统计报表",
        "filename_prefix": "地市随销统计",
        "date_label": "数据日期",
    },
    "monthly": {
        "title": "各地市随销月累计报表",
        "html_title": "各地市随销月累计报表",
        "filename_prefix": "各地市随销月累计报表",
        "date_label": "截至日期",
    },
}


def project_root() -> Path:
    return Path(__file__).resolve().parents[4]


def latest_xlsx(directory: Path, patterns: list[str] | None = None) -> Path | None:
    patterns = patterns or ["*.xlsx"]
    files = [
        path
        for pattern in patterns
        for path in directory.glob(pattern)
        if path.is_file() and not path.name.startswith("~$")
    ]
    if not files:
        return None
    return max(files, key=lambda path: path.stat().st_mtime)


def clean_header(value) -> str:
    return str(value or "").replace("\r", "").replace("\n", "").replace("\t", "").strip()


def format_fields(headers: list[str]) -> str:
    fields = [header for header in headers if header]
    return "、".join(fields) if fields else "无"


def validate_xlsx_file(path: Path) -> None:
    if path.suffix.lower() != ".xlsx":
        raise ValueError(f"{path} 文件格式不符合要求：只支持 .xlsx")
    with path.open("rb") as handle:
        signature = handle.read(4)
    if not signature.startswith(b"PK"):
        raise ValueError(f"{path} 不是标准 .xlsx 文件。请确认 BI 下载格式为 xlsx，不要把 CSV/HTML/XLS 改名为 .xlsx。")


def read_headers(path: Path, iterator) -> list[str]:
    try:
        raw_headers = next(iterator)
    except StopIteration as exc:
        raise ValueError(f"{path} 文件为空：第 1 行必须是字段名") from exc
    headers = [clean_header(value) for value in raw_headers]
    non_empty = [header for header in headers if header]
    duplicates = sorted({header for header in non_empty if non_empty.count(header) > 1})
    if duplicates:
        raise ValueError(f"{path} 第 1 行字段名重复: {', '.join(duplicates)}\n实际字段: {format_fields(headers)}")
    return headers


def require_columns(path: Path, indexes: dict[str, int], required: set[str], headers: list[str]) -> None:
    missing = sorted(required - set(indexes))
    if missing:
        raise ValueError(f"{path} 缺少字段: {', '.join(missing)}\n实际字段: {format_fields(headers)}")


def normalize_city(value) -> str:
    text = str(value or "").strip()
    return text[:-1] if text.endswith("市") else text


def decimal_value(value) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = str(value).strip().replace(",", "").replace("%", "")
    if not text:
        return Decimal("0")
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal("0")


def escape(value) -> str:
    return html.escape(str(value or ""), quote=True)


def display_label(value) -> str:
    return str(value or "").replace("(量)", "")


def parse_dates(path: Path) -> tuple[str, str]:
    match = re.search(r"_(\d{8})_(\d{6})_", path.name)
    if not match:
        return "", ""
    return match.group(1), match.group(2)


def format_day(day: str) -> str:
    if len(day) != 8:
        return day
    return f"{day[:4]}-{day[4:6]}-{day[6:8]}"


def load_rows(path: Path) -> tuple[list[str], list[dict]]:
    validate_xlsx_file(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    iterator = ws.iter_rows(values_only=True)
    try:
        headers = read_headers(path, iterator)
        indexes = {header: idx for idx, header in enumerate(headers) if header}
        require_columns(path, indexes, REQUIRED_COLUMNS, headers)

        rows: list[dict] = []
        for raw in iterator:
            if not raw or not any(value is not None and value != "" for value in raw):
                continue
            row = {header: raw[idx] if idx < len(raw) else None for header, idx in indexes.items()}
            row["地市"] = normalize_city(row.get("地市"))
            if row.get("地市"):
                rows.append(row)
    finally:
        wb.close()
    return headers, rows


def ensure_area_detail(path: Path, rows: list[dict]) -> None:
    city_names = sorted({normalize_city(row.get("地市")) for row in rows if normalize_city(row.get("地市"))})
    detail_names = [city for city in city_names if city != "全省"]
    if not detail_names:
        actual = "、".join(city_names) if city_names else "无"
        raise ValueError(f"{path} 只读取到全省汇总，没有各地市明细。请检查 BI 下载是否保留地市维度。实际地市值: {actual}")


def build_columns(headers: list[str]) -> tuple[list[dict], list[dict]]:
    fixed = [
        {"key": "地市", "label": "地市", "rowspan": 2, "kind": "text"},
        {"key": "编码", "label": "编码", "rowspan": 2, "kind": "text"},
    ]
    grouped = []
    for group, subs in GROUPS:
        children = []
        for sub in subs:
            key = f"{group}-{sub}"
            if key in headers:
                kind = "ratio" if "占比" in sub else "number"
                children.append({"key": key, "label": sub, "kind": kind})
        if children:
            grouped.append({"group": group, "children": children})

    singles = []
    for key in SINGLE_COLUMNS:
        if key in headers:
            singles.append({"key": key, "label": key, "rowspan": 2, "kind": "number"})
    return fixed + grouped, singles


def flat_metric_columns(grouped_columns: list[dict], singles: list[dict]) -> list[dict]:
    metrics = []
    for col in grouped_columns:
        if "children" in col:
            metrics.extend(col["children"])
    metrics.extend(singles)
    return metrics


def city_rows_for_rank(rows: list[dict]) -> list[tuple[int, dict]]:
    return [
        (idx, row)
        for idx, row in enumerate(rows)
        if normalize_city(row.get("地市")) != "全省"
    ]


def rank_classes(rows: list[dict], columns: list[dict]) -> dict[tuple[int, str], str]:
    classes: dict[tuple[int, str], str] = {}
    ranked_rows = city_rows_for_rank(rows)
    for col in columns:
        key = col["key"]
        values = [(idx, decimal_value(row.get(key)), normalize_city(row.get("地市"))) for idx, row in ranked_rows]
        if len(values) < HIGHLIGHT_COUNT * 2:
            continue
        unique_values = {value for _, value, _ in values}
        if len(unique_values) <= 1 or all(value == 0 for _, value, _ in values):
            continue
        top = sorted(values, key=lambda item: (-item[1], item[2]))[:HIGHLIGHT_COUNT]
        bottom = sorted(values, key=lambda item: (item[1], item[2]))[:HIGHLIGHT_COUNT]
        for idx, _, _ in bottom:
            classes[(idx, key)] = "rank-low"
        for idx, _, _ in top:
            classes[(idx, key)] = "rank-high"
    return classes


def format_number(value, kind: str) -> str:
    number = decimal_value(value)
    if kind == "ratio":
        pct = (number * Decimal("100")).quantize(Decimal("0.1"), rounding=ROUND_HALF_UP)
        return f"{pct}%"
    if number == number.to_integral_value():
        return f"{int(number):,}"
    rounded = number.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    return f"{rounded:,}"


def sort_rows(rows: list[dict]) -> list[dict]:
    province = [row for row in rows if normalize_city(row.get("地市")) == "全省"]
    cities = [row for row in rows if normalize_city(row.get("地市")) != "全省"]

    def code_sort(row: dict):
        code = str(row.get("编码") or "").strip()
        try:
            return (0, int(code), normalize_city(row.get("地市")))
        except ValueError:
            return (1, code, normalize_city(row.get("地市")))

    return province[:1] + sorted(cities, key=code_sort)


def summary_value(rows: list[dict], key: str, kind: str = "number") -> str:
    province = next((row for row in rows if normalize_city(row.get("地市")) == "全省"), None)
    if province:
        return format_number(province.get(key), kind)
    total = sum((decimal_value(row.get(key)) for row in rows), Decimal("0"))
    return format_number(total, kind)


def chart_items(rows: list[dict], key: str, reverse: bool, limit: int = 5) -> list[tuple[str, Decimal]]:
    values = [
        (normalize_city(row.get("地市")), decimal_value(row.get(key)))
        for row in rows
        if normalize_city(row.get("地市")) != "全省"
    ]
    return sorted(values, key=lambda item: ((-item[1] if reverse else item[1]), item[0]))[:limit]


def metric_values(rows: list[dict], key: str) -> list[tuple[str, Decimal]]:
    return [
        (normalize_city(row.get("地市")), decimal_value(row.get(key)))
        for row in rows
        if normalize_city(row.get("地市")) and normalize_city(row.get("地市")) != "全省"
    ]


def focus_rank_items(rows: list[dict], key: str, reverse: bool, limit: int = FOCUS_LIMIT) -> list[tuple[int, str, Decimal]]:
    values = metric_values(rows, key)
    if not values:
        return []
    sorted_values = sorted(values, key=lambda item: ((-item[1] if reverse else item[1]), item[0]))
    return [(rank, city, value) for rank, (city, value) in enumerate(sorted_values[:limit], start=1)]


def focus_rank_classes(rows: list[dict], metrics: list[dict]) -> dict[tuple[int, str], str]:
    classes: dict[tuple[int, str], str] = {}
    ranked_rows = city_rows_for_rank(rows)
    for metric in metrics:
        key = metric["key"]
        values = [(idx, decimal_value(row.get(key)), normalize_city(row.get("地市"))) for idx, row in ranked_rows]
        if not values:
            continue
        unique_values = {value for _, value, _ in values}
        if len(unique_values) <= 1:
            continue
        top = sorted(values, key=lambda item: (-item[1], item[2]))[:FOCUS_LIMIT]
        bottom = sorted(values, key=lambda item: (item[1], item[2]))[:FOCUS_LIMIT]
        for idx, _, _ in bottom:
            classes[(idx, key)] = "rank-low"
        for idx, _, _ in top:
            classes[(idx, key)] = "rank-high"
    return classes


def render_rank_rows(items: list[tuple[int, str, Decimal]], kind: str) -> str:
    if not items:
        return '<div class="rank-empty">暂无数据</div>'
    rows = []
    for rank, city, value in items:
        rows.append(
            f"""
            <div class="rank-row">
              <div class="rank-no">{rank}</div>
              <div class="rank-city">{escape(city)}</div>
              <div class="rank-value">{escape(format_number(value, kind))}</div>
            </div>
            """
        )
    return "".join(rows)


def render_focus_card(rows: list[dict], metric: dict) -> str:
    key = metric["key"]
    title = metric["title"]
    kind = metric["kind"]
    top = render_rank_rows(focus_rank_items(rows, key, reverse=True), kind)
    last = render_rank_rows(focus_rank_items(rows, key, reverse=False), kind)
    return f"""
      <section class="focus-card">
        <div class="focus-head">
          <div>
            <h2>{escape(title)}</h2>
            <div class="focus-subtitle">TOP 5 / LAST 5 地市对比</div>
          </div>
          <div class="focus-total">
            <span>全省</span>
            <strong>{summary_value(rows, key, kind)}</strong>
          </div>
        </div>
        <div class="rank-columns">
          <div class="rank-panel rank-panel-top">
            <div class="rank-title">TOP 5</div>
            {top}
          </div>
          <div class="rank-panel rank-panel-last">
            <div class="rank-title">LAST 5</div>
            {last}
          </div>
        </div>
      </section>
    """


def render_focus_sections(rows: list[dict], headers: list[str]) -> tuple[str, list[dict]]:
    available = [metric for metric in FOCUS_METRICS if metric["key"] in headers]
    missing = [metric["title"] for metric in FOCUS_METRICS if metric["key"] not in headers]
    cards = "".join(render_focus_card(rows, metric) for metric in available)
    warning = ""
    if missing:
        warning = f'<div class="missing-warning">缺少重点字段：{escape("、".join(missing))}</div>'
    return f'<div class="focus-grid">{cards}</div>{warning}', available


def render_focus_detail_table(rows: list[dict], metrics: list[dict]) -> str:
    classes = focus_rank_classes(rows, metrics)
    headers = "".join(f'<th>{escape(metric["title"])}</th>' for metric in metrics)
    body_rows = []
    for idx, row in enumerate(rows):
        city = normalize_city(row.get("地市"))
        is_province = city == "全省"
        row_class = ' class="province-row"' if is_province else ""
        cells = [f'<td class="fixed">{escape(city)}</td>']
        for metric in metrics:
            key = metric["key"]
            cell_class = "num"
            if not is_province:
                rank_class = classes.get((idx, key))
                if rank_class:
                    cell_class += f" {rank_class}"
            cells.append(f'<td class="{cell_class}">{escape(format_number(row.get(key), metric["kind"]))}</td>')
        body_rows.append(f"<tr{row_class}>{''.join(cells)}</tr>")

    return f"""
      <table class="report-table focus-table">
        <thead>
          <tr><th>地市</th>{headers}</tr>
        </thead>
        <tbody>
          {''.join(body_rows)}
        </tbody>
      </table>
    """


def render_bar_rows(items: list[tuple[str, Decimal]], max_value: Decimal, kind: str, fill_class: str) -> str:
    if not items:
        return ""
    bars = []
    for city, value in items:
        width = max(8, int((value / max_value) * Decimal("100")))
        bars.append(
            f"""
            <div class="bar-row">
              <div class="bar-label">{escape(city)}</div>
              <div class="bar-track"><div class="bar-fill {fill_class}" style="width:{width}%"></div></div>
              <div class="bar-value">{escape(format_number(value, kind))}</div>
            </div>
            """
        )
    return "".join(bars)


def render_chart(title: str, rows: list[dict], key: str, kind: str = "number") -> str:
    top = chart_items(rows, key, reverse=True)
    last = chart_items(rows, key, reverse=False)
    combined = top + last
    if not combined:
        return ""
    max_value = max((value for _, value in combined), default=Decimal("0")) or Decimal("1")
    return f"""
      <section class="chart-card">
        <h2>{escape(title)}</h2>
        <div class="chart-columns">
          <div class="chart-panel">
            <div class="chart-label chart-label-top">TOP 5</div>
            {render_bar_rows(top, max_value, kind, "bar-fill-top")}
          </div>
          <div class="chart-panel">
            <div class="chart-label chart-label-last">LAST 5</div>
            {render_bar_rows(last, max_value, kind, "bar-fill-last")}
          </div>
        </div>
      </section>
    """


def render_focus_charts(rows: list[dict], headers: list[str]) -> str:
    available = [metric for metric in FOCUS_METRICS if metric["key"] in headers]
    missing = [metric["title"] for metric in FOCUS_METRICS if metric["key"] not in headers]
    charts = "".join(render_chart(metric["title"], rows, metric["key"], metric["kind"]) for metric in available)
    warning = ""
    if missing:
        warning = f'<div class="missing-warning">缺少重点字段：{escape("、".join(missing))}</div>'
    return f"""
      <section class="chart-section">
        <div class="section-title">
          <h2>重点指标</h2>
          <span>151装维量 / FTTR-H/B装维量 / 人均价值积分，均展示 TOP 5 和 LAST 5</span>
        </div>
        <div class="charts">{charts}</div>
        {warning}
      </section>
    """


def render_table(rows: list[dict], grouped_columns: list[dict], singles: list[dict]) -> str:
    metric_columns = flat_metric_columns(grouped_columns, singles)
    classes = rank_classes(rows, metric_columns)

    top_cells = []
    second_cells = []
    for col in grouped_columns:
        if "children" in col:
            top_cells.append(f'<th class="group" colspan="{len(col["children"])}">{escape(col["group"])}</th>')
            for child in col["children"]:
                second_cells.append(f'<th>{escape(display_label(child["label"]))}</th>')
        else:
            top_cells.append(f'<th rowspan="2">{escape(display_label(col["label"]))}</th>')
    for col in singles:
        top_cells.append(f'<th rowspan="2">{escape(display_label(col["label"]))}</th>')

    body_rows = []
    for idx, row in enumerate(rows):
        is_province = normalize_city(row.get("地市")) == "全省"
        row_class = ' class="province-row"' if is_province else ""
        cells = []
        for col in grouped_columns:
            if "children" in col:
                for child in col["children"]:
                    key = child["key"]
                    cell_class = "num"
                    if not is_province:
                        rank_class = classes.get((idx, key))
                        if rank_class:
                            cell_class += f" {rank_class}"
                    cells.append(f'<td class="{cell_class}">{escape(format_number(row.get(key), child["kind"]))}</td>')
            else:
                value = row.get(col["key"])
                cells.append(f'<td class="fixed">{escape(value)}</td>')
        for col in singles:
            key = col["key"]
            cell_class = "num"
            if not is_province:
                rank_class = classes.get((idx, key))
                if rank_class:
                    cell_class += f" {rank_class}"
            cells.append(f'<td class="{cell_class}">{escape(format_number(row.get(key), col["kind"]))}</td>')
        body_rows.append(f"<tr{row_class}>{''.join(cells)}</tr>")

    return f"""
      <table class="report-table">
        <thead>
          <tr>{''.join(top_cells)}</tr>
          <tr>{''.join(second_cells)}</tr>
        </thead>
        <tbody>
          {''.join(body_rows)}
        </tbody>
      </table>
    """


def report_config(report_kind: str) -> dict:
    return REPORT_KINDS.get(report_kind, REPORT_KINDS["standard"])


def default_input_patterns(report_kind: str) -> list[str]:
    if report_kind == "monthly":
        return ["装维月累计_地市汇总_*.xlsx", "*各地市装维月累计*.xlsx"]
    return ["装维日清单_地市汇总_*.xlsx", "*.xlsx"]


def render_html(source: Path, rows: list[dict], grouped_columns: list[dict], singles: list[dict], report_kind: str = "standard") -> str:
    acct_day, month_id = parse_dates(source)
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    title_date = format_day(acct_day) if acct_day else "最新"
    config = report_config(report_kind)
    headers = ["地市"]
    for col in grouped_columns:
        if "children" in col:
            headers.extend(child["key"] for child in col["children"])
        else:
            headers.append(col["key"])
    headers.extend(col["key"] for col in singles)
    wide_table = render_table(rows, grouped_columns, singles)
    focus_charts = render_focus_charts(rows, headers)

    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{escape(config["html_title"])}_{escape(acct_day or "latest")}</title>
  <style>
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      background: #eef2f6;
      color: #172033;
      font-family: "Microsoft YaHei", "PingFang SC", Arial, sans-serif;
    }}
    .page {{
      width: 1840px;
      margin: 0 auto;
      padding: 30px 34px 34px;
      background: #f7f9fc;
    }}
    .title-row {{
      display: flex;
      align-items: flex-end;
      justify-content: space-between;
      gap: 24px;
      margin-bottom: 18px;
      border-bottom: 3px solid #1d4e89;
      padding-bottom: 16px;
    }}
    h1 {{
      margin: 0;
      color: #102a43;
      font-size: 34px;
      line-height: 1.15;
      letter-spacing: 0;
      font-weight: 800;
    }}
    .subtitle {{
      margin-top: 8px;
      font-size: 16px;
      color: #52606d;
    }}
    .meta {{
      text-align: right;
      font-size: 14px;
      color: #334e68;
      line-height: 1.8;
      white-space: nowrap;
    }}
    .focus-grid {{
      display: grid;
      grid-template-columns: repeat(3, 1fr);
      gap: 14px;
      margin-bottom: 18px;
    }}
    .focus-card {{
      background: #ffffff;
      border: 1px solid #cbd5e1;
      border-top: 5px solid #1d4e89;
      border-radius: 6px;
      padding: 16px;
      min-height: 430px;
    }}
    .focus-head {{
      display: flex;
      justify-content: space-between;
      align-items: flex-start;
      gap: 14px;
      margin-bottom: 14px;
      padding-bottom: 12px;
      border-bottom: 1px solid #d9e2ec;
    }}
    .focus-card h2 {{
      margin: 0;
      color: #102a43;
      font-size: 26px;
      line-height: 1.2;
      letter-spacing: 0;
    }}
    .focus-subtitle {{
      margin-top: 5px;
      color: #52606d;
      font-size: 14px;
    }}
    .focus-total {{
      min-width: 116px;
      text-align: right;
    }}
    .focus-total span {{
      display: block;
      color: #52606d;
      font-size: 13px;
      margin-bottom: 5px;
    }}
    .focus-total strong {{
      display: block;
      color: #102a43;
      font-size: 26px;
      line-height: 1.1;
      font-variant-numeric: tabular-nums;
    }}
    .rank-columns {{
      display: grid;
      grid-template-columns: 1fr 1fr;
      gap: 12px;
    }}
    .rank-panel {{
      border: 1px solid #d9e2ec;
      border-radius: 6px;
      overflow: hidden;
      background: #f8fafc;
    }}
    .rank-title {{
      padding: 9px 10px;
      color: #ffffff;
      font-size: 15px;
      font-weight: 800;
      letter-spacing: 0;
    }}
    .rank-panel-top .rank-title {{
      background: #137333;
    }}
    .rank-panel-last .rank-title {{
      background: #b42318;
    }}
    .rank-row {{
      display: grid;
      grid-template-columns: 34px 1fr 88px;
      align-items: center;
      min-height: 44px;
      padding: 8px 9px;
      border-top: 1px solid #e2e8f0;
      background: #ffffff;
      gap: 8px;
      font-size: 15px;
    }}
    .rank-no {{
      width: 26px;
      height: 26px;
      line-height: 26px;
      border-radius: 50%;
      background: #e2e8f0;
      color: #172033;
      text-align: center;
      font-weight: 800;
      font-variant-numeric: tabular-nums;
    }}
    .rank-city {{
      color: #172033;
      font-weight: 800;
      white-space: nowrap;
      overflow: hidden;
      text-overflow: ellipsis;
    }}
    .rank-value {{
      color: #102a43;
      font-weight: 900;
      text-align: right;
      font-variant-numeric: tabular-nums;
    }}
    .rank-empty {{
      padding: 20px;
      color: #52606d;
      background: #ffffff;
      font-size: 14px;
    }}
    .missing-warning {{
      margin: -6px 0 14px;
      padding: 10px 12px;
      border: 1px solid #f8b4a6;
      border-radius: 6px;
      background: #fff4f2;
      color: #b42318;
      font-size: 14px;
      font-weight: 700;
    }}
    .table-wrap {{
      background: #ffffff;
      border: 1px solid #cbd5e1;
      border-radius: 6px;
      overflow: hidden;
      margin-bottom: 16px;
    }}
    .report-table {{
      width: 100%;
      border-collapse: collapse;
      table-layout: fixed;
      font-size: 16px;
      line-height: 1.25;
    }}
    .report-table th {{
      border: 1px solid #94a3b8;
      background: #dbeafe;
      color: #0f2747;
      padding: 10px 6px;
      text-align: center;
      font-weight: 800;
      vertical-align: middle;
    }}
    .report-table th.group {{
      background: #bfdbfe;
      font-size: 12px;
    }}
    .report-table td {{
      border: 1px solid #cbd5e1;
      padding: 9px 6px;
      text-align: center;
      background: #ffffff;
      vertical-align: middle;
      white-space: nowrap;
    }}
    .report-table td.fixed {{
      font-weight: 700;
      color: #172033;
    }}
    .report-table td.num {{
      text-align: center;
      font-variant-numeric: tabular-nums;
    }}
    .report-table tbody tr:nth-child(even) td {{
      background: #f8fafc;
    }}
    .report-table .province-row td {{
      background: #1d4e89 !important;
      color: #ffffff;
      font-weight: 800;
    }}
    .report-table td.rank-high {{
      background: #ecfdf3 !important;
      color: #116329;
      font-weight: 800;
    }}
    .report-table td.rank-low {{
      background: #fff1f0 !important;
      color: #b42318;
      font-weight: 800;
    }}
    .charts {{
      display: grid;
      grid-template-columns: repeat(3, 1fr);
      gap: 12px;
    }}
    .chart-section {{
      margin-top: 16px;
    }}
    .section-title {{
      display: flex;
      align-items: baseline;
      justify-content: space-between;
      gap: 16px;
      margin: 0 0 10px;
    }}
    .section-title h2 {{
      margin: 0;
      color: #102a43;
      font-size: 22px;
      line-height: 1.2;
    }}
    .section-title span {{
      color: #52606d;
      font-size: 14px;
    }}
    .chart-card {{
      background: #ffffff;
      border: 1px solid #d9e2ec;
      border-radius: 6px;
      padding: 12px 12px 14px;
    }}
    .chart-card h2 {{
      margin: 0 0 12px;
      color: #102a43;
      font-size: 16px;
      line-height: 1.2;
    }}
    .chart-columns {{
      display: grid;
      grid-template-columns: 1fr 1fr;
      gap: 10px;
    }}
    .chart-label {{
      margin-bottom: 7px;
      font-size: 13px;
      line-height: 1.2;
      font-weight: 900;
    }}
    .chart-label-top {{
      color: #116329;
    }}
    .chart-label-last {{
      color: #b42318;
    }}
    .bar-row {{
      display: grid;
      grid-template-columns: 44px 1fr 46px;
      align-items: center;
      gap: 6px;
      margin: 7px 0;
      font-size: 12px;
    }}
    .bar-label {{
      color: #334e68;
      font-weight: 700;
    }}
    .bar-track {{
      height: 12px;
      background: #e2e8f0;
      border-radius: 999px;
      overflow: hidden;
    }}
    .bar-fill {{
      height: 100%;
      border-radius: 999px;
    }}
    .bar-fill-top {{
      background: linear-gradient(90deg, #1d4e89, #38bdf8);
    }}
    .bar-fill-last {{
      background: linear-gradient(90deg, #b42318, #f97316);
    }}
    .bar-value {{
      color: #102a43;
      font-weight: 800;
      text-align: right;
      font-variant-numeric: tabular-nums;
    }}
    .footnote {{
      margin-top: 12px;
      color: #52606d;
      font-size: 13px;
      line-height: 1.6;
    }}
  </style>
</head>
<body>
  <main class="page">
    <header class="title-row">
      <div>
        <h1>{escape(config["title"])}</h1>
        <div class="subtitle">{escape(config["date_label"])}：{escape(title_date)}　账期：{escape(month_id or "-")}</div>
      </div>
      <div class="meta">
        <div>生成时间：{escape(generated_at)}</div>
        <div>数据来源：{escape(source.name)}</div>
      </div>
    </header>

    <section class="table-wrap">
      {wide_table}
    </section>

    {focus_charts}

    <div class="footnote">
      上方宽表展示各地市全量业务指标；下方柱状图展示 151装维量、FTTR-H/B装维量、人均价值积分 TOP 5 和 LAST 5。“全省”行不参与地市排名。
    </div>
  </main>
</body>
</html>
"""


def default_output_path(root: Path, source: Path, report_kind: str = "standard") -> Path:
    acct_day, _ = parse_dates(source)
    suffix = acct_day or datetime.now().strftime("%Y%m%d")
    prefix = report_config(report_kind)["filename_prefix"]
    return root / "output" / f"{prefix}_{suffix}.html"


def main() -> None:
    parser = argparse.ArgumentParser(description="生成地市随销统计 HTML 报表")
    parser.add_argument("--input", help="地市汇总 Excel；默认读取 temp/data/area 本轮下载 xlsx")
    parser.add_argument("--area-dir", default="temp/data/area")
    parser.add_argument("--output", help="HTML 输出路径；默认 output/地市随销统计_{acct_day}.html")
    parser.add_argument(
        "--report-kind",
        choices=sorted(REPORT_KINDS),
        default="standard",
        help="报表类型：standard=地市随销统计报表，monthly=各地市随销月累计报表",
    )
    args = parser.parse_args()

    root = project_root()
    if args.input:
        source = Path(args.input)
        if not source.is_absolute():
            source = root / source
    else:
        source = latest_xlsx(root / args.area_dir, default_input_patterns(args.report_kind))
        if not source:
            raise SystemExit(f"未找到地市汇总 Excel: {root / args.area_dir}")

    headers, rows = load_rows(source)
    ensure_area_detail(source, rows)
    rows = sort_rows(rows)
    grouped_columns, singles = build_columns(headers)
    if not grouped_columns:
        raise SystemExit("未找到可用于生成报表的指标列。")

    if args.output:
        output = Path(args.output)
        if not output.is_absolute():
            output = root / output
    else:
        output = default_output_path(root, source, args.report_kind)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(
        render_html(source, rows, grouped_columns, singles, args.report_kind),
        encoding="utf-8",
    )

    print(f"=== {report_config(args.report_kind)['title']} ===")
    print(f"来源文件: {source}")
    print(f"输出文件: {output}")
    print(f"数据行数: {len(rows)}")


if __name__ == "__main__":
    main()
