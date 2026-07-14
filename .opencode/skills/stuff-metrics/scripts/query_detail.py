#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
import sys
import warnings
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError as exc:
    raise SystemExit("缺少依赖 openpyxl，请使用 uv --with openpyxl 运行。") from exc

warnings.filterwarnings("ignore", message="Workbook contains no default style.*", category=UserWarning)

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")


STAFF_TYPE_ALIASES = {
    "official": ["official", "正式", "正式人员"],
    "intern": ["intern", "实习", "实习人员"],
}
DATASETS = {
    "official": {
        "label": "正式人员",
        "source": "装维随销发展日清单（正式人员）",
        "dir": "temp/data/official",
    },
    "intern": {
        "label": "实习人员",
        "source": "装维随销发展日清单（实习人员）",
        "dir": "temp/data/intern",
    },
}
DEFAULT_MAX_OUTPUT_ROWS = 20
OPTIONAL_INTERN_COLUMNS = ["实习开始时间"]
REQUIRED_COLUMNS = {"地市", "区县", "装维姓名", "工号"}


def project_root() -> Path:
    return Path(__file__).resolve().parents[4]


def normalize_staff_type(value: str) -> str:
    text = str(value or "").strip().lower()
    for normalized, aliases in STAFF_TYPE_ALIASES.items():
        if text in aliases:
            return normalized
    raise ValueError("人员类型只能是 official/intern，或 正式/实习")


def normalize_city(value: str) -> str:
    text = str(value or "").strip()
    if text.endswith("市"):
        text = text[:-1]
    return text


def text_match(value, query: str, normalize=None) -> bool:
    value_text = str(value or "").strip()
    query_text = str(query or "").strip()
    if normalize:
        value_text = normalize(value_text)
        query_text = normalize(query_text)
    return bool(value_text and query_text and (value_text == query_text or query_text in value_text or value_text in query_text))


def latest_xlsx(directory: Path) -> Path | None:
    files = [
        path
        for path in directory.glob("*.xlsx")
        if path.is_file() and not path.name.startswith("~$")
    ]
    if not files:
        return None
    return max(files, key=lambda path: path.stat().st_mtime)


def clean_header(value) -> str:
    return str(value or "").replace("\r", "").replace("\n", "").replace("\t", "").strip()


def format_fields(headers: list[str]) -> str:
    fields = [header for header in headers if header]
    return "、".join(fields) if fields else "无"


def validate_xlsx_file(path: Path) -> None:
    if path.suffix.lower() != ".xlsx":
        raise ValueError(f"{path} 文件格式不符合要求：只支持 .xlsx")
    with path.open("rb") as handle:
        signature = handle.read(4)
    if not signature.startswith(b"PK"):
        raise ValueError(f"{path} 不是标准 .xlsx 文件。请确认 BI 下载格式为 xlsx，不要把 CSV/HTML/XLS 改名为 .xlsx。")


def read_headers(path: Path, iterator) -> list[str]:
    try:
        raw_headers = next(iterator)
    except StopIteration as exc:
        raise ValueError(f"{path} 文件为空：第 1 行必须是字段名") from exc
    headers = [clean_header(value) for value in raw_headers]
    non_empty = [header for header in headers if header]
    duplicates = sorted({header for header in non_empty if non_empty.count(header) > 1})
    if duplicates:
        raise ValueError(f"{path} 第 1 行字段名重复: {', '.join(duplicates)}\n实际字段: {format_fields(headers)}")
    return headers


def require_columns(path: Path, indexes: dict[str, int], required: set[str], headers: list[str]) -> None:
    missing = sorted(required - set(indexes))
    if missing:
        raise ValueError(f"{path} 缺少字段: {', '.join(missing)}\n实际字段: {format_fields(headers)}")


def load_matching_rows(path: Path, staff_label: str, city: str | None, county: str | None, person: str | None):
    validate_xlsx_file(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    iterator = ws.iter_rows(values_only=True)
    try:
        headers = read_headers(path, iterator)
        indexes = {header: idx for idx, header in enumerate(headers) if header}
        require_columns(path, indexes, REQUIRED_COLUMNS, headers)

        matched = []
        for row in iterator:
            if city and not text_match(row[indexes["地市"]], city, normalize_city):
                continue
            if county and not text_match(row[indexes["区县"]], county):
                continue
            if person:
                by_name = text_match(row[indexes["装维姓名"]], person)
                by_code = text_match(row[indexes["工号"]], person)
                if not (by_name or by_code):
                    continue
            matched.append([staff_label, *list(row[: len(headers)])])
    finally:
        wb.close()
    return ["人员类型", *headers], matched


def markdown_table(headers: list[str], rows: list[list]) -> str:
    lines = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join("---" for _ in headers) + " |",
    ]
    for row in rows:
        lines.append("| " + " | ".join(str(value or "") for value in row[: len(headers)]) + " |")
    return "\n".join(lines)


def print_table_or_limit_notice(headers: list[str], rows: list[list], max_rows: int) -> None:
    if len(rows) > max_rows:
        print(f"明细行数超过 {max_rows} 行，不在消息栏罗列；请打开 Excel 查看完整清单。")
        return
    print(markdown_table(headers, rows))


def write_excel(output_path: Path, headers: list[str], rows: list[list], meta: dict) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "人员明细"
    ws.append(headers)
    for row in rows:
        ws.append(row)

    meta_ws = wb.create_sheet("来源")
    meta_ws.append(["项目", "值"])
    meta_ws.append(["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    for key, value in meta.items():
        meta_ws.append([key, value])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def safe_save_excel(output_path: Path, headers: list[str], rows: list[list], meta: dict) -> Path:
    try:
        write_excel(output_path, headers, rows, meta)
        return output_path
    except PermissionError:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        fallback = output_path.with_name(f"{output_path.stem}_{timestamp}{output_path.suffix}")
        write_excel(fallback, headers, rows, meta)
        return fallback


def main() -> None:
    parser = argparse.ArgumentParser(description="读取装维随销发展日清单范围明细")
    parser.add_argument("--staff-type", required=True, help="人员类型: official/intern，或 正式/实习")
    parser.add_argument("--city", help="地市名称")
    parser.add_argument("--county", help="区县名称")
    parser.add_argument("--person", help="人员姓名或工号")
    parser.add_argument("--official-dir", default="temp/data/official")
    parser.add_argument("--intern-dir", default="temp/data/intern")
    parser.add_argument("--output", help="可选，输出 Excel 路径")
    parser.add_argument("--max-output", type=int, default=DEFAULT_MAX_OUTPUT_ROWS, help="消息栏最多输出行数")
    args = parser.parse_args()

    scope_count = sum(1 for value in [args.city, args.county, args.person] if value)
    if scope_count > 1:
        raise SystemExit("范围参数最多指定一个：--city、--county 或 --person。")

    root = project_root()
    staff_type = normalize_staff_type(args.staff_type)
    DATASETS["official"]["dir"] = args.official_dir
    DATASETS["intern"]["dir"] = args.intern_dir
    dataset = DATASETS[staff_type]
    source_path = latest_xlsx(root / dataset["dir"])
    if not source_path:
        raise SystemExit(f"未找到{dataset['label']} Excel: {root / dataset['dir']}")

    headers, rows = load_matching_rows(source_path, dataset["label"], args.city, args.county, args.person)
    if not rows:
        raise SystemExit("未找到匹配的人员明细。")

    meta = {
        "人员类型": dataset["label"],
        "来源文件": str(source_path),
        "匹配行数": len(rows),
        "字段列表": ", ".join(headers),
    }
    if staff_type == "intern":
        intern_fields = [column for column in OPTIONAL_INTERN_COLUMNS if column in headers]
        meta["实习扩展字段"] = ", ".join(intern_fields) if intern_fields else "当前文件未包含"

    if scope_count == 0:
        output = root / (args.output or f"output/province_detail_{dataset['label']}.xlsx")
        saved = safe_save_excel(output, headers, rows, {**meta, "范围": "全省"})
        print(f"全省人员明细已生成: {saved}")
        print(f"明细行数: {len(rows)}")
        print(f"来源文件: {source_path}")
        if len(rows) > args.max_output:
            print(f"明细行数超过 {args.max_output} 行，不在消息栏罗列；请打开 Excel 查看完整清单。")
        return

    if args.city:
        city_name = normalize_city(args.city)
        output = root / (args.output or f"output/city_detail_{city_name}_{dataset['label']}.xlsx")
        saved = safe_save_excel(output, headers, rows, {**meta, "地市": args.city})
        print(f"地市人员明细已生成: {saved}")
        print(f"明细行数: {len(rows)}")
        if len(rows) > args.max_output:
            print(f"明细行数超过 {args.max_output} 行，不在消息栏罗列；请打开 Excel 查看完整清单。")
        return

    if args.county:
        output = root / (args.output or f"output/county_detail_{args.county}_{dataset['label']}.xlsx")
        saved = safe_save_excel(output, headers, rows, {**meta, "区县": args.county})
        print(f"区县人员明细已生成: {saved}")
        print(f"明细行数: {len(rows)}")
        print_table_or_limit_notice(headers, rows, args.max_output)
        return

    if len(rows) > args.max_output:
        print(f"匹配人员记录 {len(rows)} 行，超过 {args.max_output} 行，不在消息栏罗列；请提供更精确的姓名或工号。")
        print(f"来源文件: {source_path}")
        return
    print(markdown_table(headers, rows))


if __name__ == "__main__":
    main()
