#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# /// script
# dependencies = [
#   "requests>=2.28",
#   "urllib3>=1.26",
#   "openpyxl>=3.1",
#   "python-dateutil>=2.8",
# ]
# requires-python = ">=3.10"
# ///

"""
BI系统数据下载脚本
功能：
  1. 线上佣金出账月清单 - 登录BI -> 读取CSV -> 批量下载代理商月清单数据
  2. 电渠监控日表 - 登录BI -> 下载指定日期的监控数据
  3. 数据报告 - 登录BI -> 下载指定报告的图表数据
  4. 智能参数发现 - 自动从API错误响应中提取缺失参数

安全设计：所有凭证输入在终端直接完成，不会暴露给 Claude
"""

import argparse
import csv
import json
import os
import re
import sys
import time
import urllib.parse
import xml.etree.ElementTree as ET
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from getpass import getpass

import requests
import urllib3
import openpyxl
from openpyxl import Workbook

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

CREDENTIALS_FILE = ".bi_credentials"
SESSION = requests.Session()

REPORT_PATHS = {
    "commission": "省公司数据集/全渠/线上佣金稽核/线上佣金出账月清单_副本",
    "daily": "zxt-test/0319电渠监控日表（日常筛选测试）",
    "sales_monthly": "zxt-test/250830/销售点月报-下载专用",
}

HR_DATASETS = {
    "channel_staff": {
        "path": "省公司数据集/人力资源部/人效分析/结果展示类/渠道三大岗位人员清单",
        "name": "渠道三大岗位人员清单",
    },
    "manager_eval": {
        "path": "省公司数据集/人力资源部/人效分析/结果展示类/渠道经理人员评价表",
        "name": "渠道经理人员评价表",
    },
    "direct_sales_eval": {
        "path": "省公司数据集/人力资源部/人效分析/结果展示类/公众渠道直销经理人员评价表",
        "name": "公众渠道直销经理人员评价表",
    },
    "sales_rep_eval": {
        "path": "省公司数据集/人力资源部/人效分析/结果展示类/营业代表人员评价表",
        "name": "营业代表人员评价表",
    },
}

HR_REPORT_PATH = "省公司报表/人力资源部/人员效能分析/渠道相关"

HR_CHARTS = {
    "channel_staff": "图表4",
    "manager_eval": "图表3",
    "direct_sales_eval": "图表1",
    "sales_rep_eval": "图表6",
}

API_BASE = "https://137.0.247.223:31004/bi/api"
REPORT_API_BASE = "https://137.0.245.223:31004/sdk/bi/api"

QRY_TYPES = [
    "sql",
    "mongo",
    "excel",
    "composite",
    "cloud",
    "embed",
    "custom",
    "data_flow",
    "restful",
]
REPORT_TYPES = ["db", "dblink", "adb", "vividdb", "portalCell", "vividdblink"]
PARAM_TYPES = [
    "string",
    "boolean",
    "float",
    "double",
    "char",
    "byte",
    "short",
    "int",
    "long",
    "date",
    "time",
    "dateTime",
]


class Colors:
    GREEN = "\033[0;32m"
    YELLOW = "\033[1;33m"
    RED = "\033[0;31m"
    BLUE = "\033[0;34m"
    NC = "\033[0m"

    @classmethod
    def print(cls, color: str, message: str):
        print(f"{color}{message}{cls.NC}")

    @classmethod
    def success(cls, message: str):
        cls.print(cls.GREEN, message)

    @classmethod
    def warning(cls, message: str):
        cls.print(cls.YELLOW, message)

    @classmethod
    def error(cls, message: str):
        cls.print(cls.RED, message)

    @classmethod
    def info(cls, message: str):
        cls.print(cls.BLUE, message)


def load_credentials() -> tuple:
    env_user = (os.getenv("BI_USER") or os.getenv("FSA_BI_USER") or "").strip()
    env_pass = (os.getenv("BI_PASS") or os.getenv("FSA_BI_PASS") or "").strip()
    if env_user and env_pass:
        return (env_user, env_pass)

    try:
        from keyring_manager import load_credentials_keyring, load_credentials_interactive, save_credentials_keyring
        username, password = load_credentials_keyring()
        if username and password:
            return (username, password)
    except ImportError:
        pass

    # 云端无 Secret 配置入口时，可在运行环境创建未提交的凭证文件。
    credential_files = [
        os.getenv("FSA_BI_CREDENTIALS_FILE"),
        ".runtime/.bi_credentials",
        ".secrets/bi_credentials",
        ".bi_credentials",
    ]
    for creds_file in [path for path in credential_files if path]:
        if not os.path.exists(creds_file):
            continue
        try:
            with open(creds_file, "r", encoding="utf-8-sig") as f:
                creds = {}
                for line in f:
                    line = line.strip()
                    if line and "=" in line and not line.startswith("#"):
                        k, v = line.split("=", 1)
                        creds[k.strip()] = v.strip()
                if creds.get("bi_user") and creds.get("bi_pass"):
                    return (creds["bi_user"], creds["bi_pass"])
        except Exception as e:
            print(f"[DEBUG] read creds file error: {e}", file=sys.stderr)

    try:
        from keyring_manager import load_credentials_interactive, save_credentials_keyring
        print("\n[提示] 未在密钥链中找到BI凭证，请输入")
        username, password = load_credentials_interactive()
        if username and password:
            save_credentials_keyring(username, password)
            print("[提示] 凭证已保存到密钥链")
            return (username, password)
    except ImportError:
        pass

    return (None, None)


def get_last_month() -> str:
    today = datetime.today()
    last_month = today - relativedelta(months=1)
    return last_month.strftime("%Y%m")


def get_yesterday() -> str:
    yesterday = datetime.today() - timedelta(days=1)
    return yesterday.strftime("%Y%m%d")


def get_param_value(params: dict, name: str) -> str | None:
    value = params.get(name)
    if isinstance(value, tuple):
        value = value[0]
    return str(value) if value else None


def validate_field_service_params(config_name: str, params: dict) -> None:
    if not config_name.startswith("field-service-agent-"):
        return

    acct_day = get_param_value(params, "acct_day")
    month_id = get_param_value(params, "month_id")
    errors = []

    if not acct_day or not re.fullmatch(r"\d{8}", acct_day):
        errors.append("acct_day 必须提供 8 位日期，例如 20260706")
    if not month_id or not re.fullmatch(r"\d{6}", month_id):
        errors.append("month_id 必须提供 6 位月份，例如 202607")

    acct_date = None
    if acct_day and re.fullmatch(r"\d{8}", acct_day):
        try:
            acct_date = datetime.strptime(acct_day, "%Y%m%d").date()
        except ValueError:
            errors.append(f"acct_day 不是有效日期: {acct_day}")

    if acct_day and month_id and re.fullmatch(r"\d{8}", acct_day) and re.fullmatch(r"\d{6}", month_id):
        if month_id != acct_day[:6]:
            errors.append(f"month_id 必须等于 acct_day 前 6 位: month_id={month_id}, acct_day={acct_day}")

    if acct_date and acct_date > datetime.today().date():
        errors.append(f"acct_day 不能是未来日期: {acct_day}")

    if errors:
        raise ValueError("日期参数错误: " + "；".join(errors))


def build_output_tokens(params: dict) -> dict[str, str]:
    run_time = datetime.now()
    timestamp = run_time.strftime("%Y%m%d%H%M%S")
    acct_day = (
        get_param_value(params, "acct_day")
        or get_param_value(params, "p_acct_day")
        or get_param_value(params, "acc_day")
    )
    month_id = get_param_value(params, "month_id")

    if not acct_day or not re.fullmatch(r"\d{8}", acct_day):
        acct_day = run_time.strftime("%Y%m%d")
    if not month_id or not re.fullmatch(r"\d{6}", month_id):
        month_id = acct_day[:6]

    return {
        "date": f"{acct_day}_{month_id}_{timestamp}",
        "acct_day": acct_day,
        "month_id": month_id,
        "timestamp": timestamp,
    }


def render_output_template(template: str, ext: str, params: dict) -> str:
    tokens = build_output_tokens(params)
    tokens["ext"] = ext
    filename = template
    for key, value in tokens.items():
        filename = filename.replace(f"{{{key}}}", value)
    return filename


def read_csv_agents(csv_file: str) -> list:
    agents = []
    with open(csv_file, "r", encoding="utf-8") as f:
        reader = csv.reader(f)
        for row in reader:
            if len(row) < 3:
                continue
            org_id, operators_nbr, operators_name = row[:3]
            if org_id == "ORG_ID":
                continue
            if not operators_nbr.strip():
                continue
            agents.append((operators_nbr.strip(), operators_name.strip()))
    return agents


def login(username: str, password: str, api_base: str = None) -> str:
    encoded_password = urllib.parse.quote(password, safe="")
    base_url = api_base if api_base else API_BASE
    url = f"{base_url}?action=login&adminv={username}&passv={encoded_password}"

    try:
        response = SESSION.post(url, verify=False, timeout=30)
        content = response.content

        if len(content) > 4 and (
            content[:2] == b"PK" or content[:4] == b"\xd0\xcf\x11\xe0"
        ):
            raise Exception("登录返回了文件而非token")

        text = content.decode("utf-8", errors="ignore")

        try:
            root = ET.fromstring(text)
            message_elem = root.find(".//message")
            if message_elem is not None and message_elem.text:
                return message_elem.text.strip()
        except:
            pass

        token_match = re.search(r"<message>([^<]+)</message>", text)
        if token_match:
            return token_match.group(1)

        raise Exception(f"登录失败：{text[:200]}")

    except requests.exceptions.RequestException as e:
        raise Exception(f"网络错误：{e}")


def download_commission_data(
    token: str, month: str, agent_code: str, agent_name: str, output_dir: str
) -> tuple:
    report_path = REPORT_PATHS["commission"]

    xml_data = f"""<?xml version="1.0" encoding="UTF-8"?>
<info>
    <qryPath>{report_path}</qryPath>
    <qryType>sql</qryType>
    <fileType>xlsx</fileType>
    <params>
        <param><name>月账期</name><type>string</type><value>{month}</value></param>
        <param><name>代理商编码</name><type>string</type><value>{agent_code}</value></param>
    </params>
</info>"""

    url = f"{API_BASE}?action=downloadQuery&token={token}"

    try:
        response = SESSION.post(
            url,
            headers={"Content-Type": "application/x-www-form-urlencoded"},
            data={"xmlData": xml_data},
            verify=False,
            timeout=120,
        )

        output_path = os.path.join(output_dir, f"{agent_name}.xlsx")
        content = response.content

        if len(content) > 4 and (
            content[:2] == b"PK" or content[:4] == b"\xd0\xcf\x11\xe0"
        ):
            with open(output_path, "wb") as f:
                f.write(content)
            file_size = os.path.getsize(output_path)
            file_size_str = (
                f"{file_size / 1024:.1f}KB"
                if file_size < 1024 * 1024
                else f"{file_size / 1024 / 1024:.1f}MB"
            )
            Colors.success(f"   [OK] {agent_name}.xlsx ({file_size_str})")
            return (True, False)
        else:
            try:
                error_text = content.decode("utf-8", errors="ignore")
                if "token 过期" in error_text or "token expired" in error_text.lower():
                    Colors.warning(f"   [WARN] {agent_name}: Token已过期")
                    return (False, True)
                Colors.error(f"   [FAIL] {agent_name}: 下载失败 - {error_text[:200]}")
            except:
                Colors.error(f"   [FAIL] {agent_name}: 下载失败")
            return (False, False)

    except requests.exceptions.RequestException as e:
        Colors.error(f"   [FAIL] {agent_name}: 网络错误 - {e}")
        return (False, False)


def download_daily_data(token: str, date: str, output_dir: str) -> tuple:
    report_path = REPORT_PATHS["daily"]

    xml_data = f"""<?xml version="1.0" encoding="UTF-8"?>
<info>
    <qryPath>{report_path}</qryPath>
    <qryType>sql</qryType>
    <fileType>xlsx</fileType>
    <params>
        <param><name>日账期</name><type>string</type><value>{date}</value></param>
    </params>
</info>"""

    url = f"{API_BASE}?action=downloadQuery&token={token}"

    try:
        response = SESSION.post(
            url,
            headers={"Content-Type": "application/x-www-form-urlencoded"},
            data={"xmlData": xml_data},
            verify=False,
            timeout=120,
        )

        output_path = os.path.join(output_dir, f"电渠监控日表_{date}.xlsx")
        content = response.content

        if len(content) > 4 and (
            content[:2] == b"PK" or content[:4] == b"\xd0\xcf\x11\xe0"
        ):
            with open(output_path, "wb") as f:
                f.write(content)
            file_size = os.path.getsize(output_path)
            file_size_str = (
                f"{file_size / 1024:.1f}KB"
                if file_size < 1024 * 1024
                else f"{file_size / 1024 / 1024:.1f}MB"
            )
            Colors.success(f"   [OK] 电渠监控日表_{date}.xlsx ({file_size_str})")
            return (True, False)
        else:
            try:
                error_text = content.decode("utf-8", errors="ignore")
                if "token 过期" in error_text or "token expired" in error_text.lower():
                    Colors.warning(f"   [WARN] Token已过期")
                    return (False, True)
                Colors.error(f"   [FAIL] 下载失败 - {error_text[:200]}")
            except:
                Colors.error(f"   [FAIL] 下载失败")
            return (False, False)

    except requests.exceptions.RequestException as e:
        Colors.error(f"   [FAIL] 网络错误 - {e}")
        return (False, False)


def download_sales_monthly_data(token: str, month: str, output_dir: str) -> tuple:
    report_path = REPORT_PATHS["sales_monthly"]

    xml_data = f"""<?xml version="1.0" encoding="UTF-8"?>
<info>
    <qryPath>{report_path}</qryPath>
    <qryType>sql</qryType>
    <fileType>xlsx</fileType>
    <params>
        <param><name>acct_month</name><type>string</type><value>{month}</value></param>
    </params>
</info>"""

    url = f"{API_BASE}?action=downloadQuery&token={token}"

    try:
        try:
            from download_validator import (
                AbnormalDataPredictor,
                require_user_confirmation_for_anomaly,
            )
        except ImportError:
            pass

        response = SESSION.post(
            url,
            headers={"Content-Type": "application/x-www-form-urlencoded"},
            data={"xmlData": xml_data},
            verify=False,
            timeout=120,
        )

        output_path = os.path.join(
            output_dir, f"{month}销售点月报_新集群-基础表格.xlsx"
        )
        content = response.content

        if len(content) > 4 and (
            content[:2] == b"PK" or content[:4] == b"\xd0\xcf\x11\xe0"
        ):
            with open(output_path, "wb") as f:
                f.write(content)

            try:
                import pandas as pd
                df_temp = pd.read_excel(output_path, nrows=10)
                row_count = len(df_temp)

                import datetime as dt
                current_hour = dt.datetime.now().hour
                is_anomaly, _ = AbnormalDataPredictor.check_row_count_anomaly(
                    row_count, current_hour
                )

                if is_anomaly:
                    confirmed = require_user_confirmation_for_anomaly(
                        row_count, df_temp,
                        message=f"检测到数据异常（当前账期{month}），是否继续使用?"
                    )
                    if not confirmed:
                        os.remove(output_path)
                        Colors.warning("[取消] 用户取消使用异常数据")
                        return (False, False)
            except Exception:
                pass

            file_size = os.path.getsize(output_path)
            file_size_str = (
                f"{file_size / 1024:.1f}KB"
                if file_size < 1024 * 1024
                else f"{file_size / 1024 / 1024:.1f}MB"
            )
            Colors.success(
                f"   [OK] {month}销售点月报_新集群-基础表格.xlsx ({file_size_str})"
            )
            return (True, False)
        else:
            try:
                error_text = content.decode("utf-8", errors="ignore")
                if "token 过期" in error_text or "token expired" in error_text.lower():
                    Colors.warning(f"   [WARN] Token已过期")
                    return (False, True)
                Colors.error(f"   [FAIL] 下载失败 - {error_text[:200]}")
            except:
                Colors.error(f"   [FAIL] 下载失败")
            return (False, False)

    except requests.exceptions.RequestException as e:
        Colors.error(f"   [FAIL] 网络错误 - {e}")
        return (False, False)


def download_dataflow_data(
    token: str,
    report_path: str,
    output_name: str,
    output_dir: str,
    qry_type: str = "data_flow",
) -> tuple:
    xml_data = f"""<?xml version="1.0" encoding="UTF-8"?>
<info>
    <qryPath>{report_path}</qryPath>
    <qryType>{qry_type}</qryType>
    <fileType>xlsx</fileType>
</info>"""

    url = f"{API_BASE}?action=downloadQuery&token={token}"

    try:
        response = SESSION.post(
            url,
            headers={"Content-Type": "application/x-www-form-urlencoded"},
            data={"xmlData": xml_data},
            verify=False,
            timeout=120,
        )

        output_path = os.path.join(output_dir, f"{output_name}.xlsx")
        content = response.content

        if len(content) > 4 and (
            content[:2] == b"PK" or content[:4] == b"\xd0\xcf\x11\xe0"
        ):
            with open(output_path, "wb") as f:
                f.write(content)
            file_size = os.path.getsize(output_path)
            file_size_str = (
                f"{file_size / 1024:.1f}KB"
                if file_size < 1024 * 1024
                else f"{file_size / 1024 / 1024:.1f}MB"
            )
            Colors.success(f"   [OK] {output_name}.xlsx ({file_size_str})")
            return (True, False)
        else:
            try:
                error_text = content.decode("utf-8", errors="ignore")
                if "token 过期" in error_text or "token expired" in error_text.lower():
                    Colors.warning(f"   [WARN] Token已过期")
                    return (False, True)
                Colors.error(f"   [FAIL] 下载失败 - {error_text[:200]}")
            except:
                Colors.error(f"   [FAIL] 下载失败")
            return (False, False)

    except requests.exceptions.RequestException as e:
        Colors.error(f"   [FAIL] 网络错误 - {e}")
        return (False, False)


def download_report_data(
    token: str, dbpath: str, ename: str, output_dir: str, params: dict = None
) -> tuple:
    page_size = 10000
    page_now = 1

    all_headers = []
    all_rows = []
    total_rows = 0

    # Build params XML if params provided
    params_xml = ""
    if params:
        for k, v in params.items():
            params_xml += (
                f"<param><name>{k}</name><type>string</type><value>{v}</value></param>"
            )

    while True:
        xml_data = f"""<?xml version="1.0" encoding="UTF-8"?>
<info>
    <page>
        <pageSize>{page_size}</pageSize>
        <pageNow>{page_now}</pageNow>
    </page>
    <getdata>
        <dbpath>{dbpath}</dbpath>
        <ename>{ename}</ename>
        {params_xml}
    </getdata>
</info>"""

        url = f"{REPORT_API_BASE}?action=getElemData&token={token}"

        try:
            response = SESSION.post(
                url,
                headers={"Content-Type": "application/x-www-form-urlencoded"},
                data={"xmlData": xml_data},
                verify=False,
                timeout=120,
            )

            content = response.content.decode("utf-8", errors="ignore")

            if content.strip().startswith("<?xml"):
                try:
                    root = ET.fromstring(content)

                    page_elem = root.find(".//page")
                    if page_elem is not None:
                        row_count_elem = page_elem.find("rowCount")
                        if row_count_elem is not None:
                            total_rows = int(row_count_elem.text)

                    grid_elem = root.find(".//grid")
                    if grid_elem is None:
                        Colors.error(f"   [FAIL] 未找到数据")
                        return (False, False)

                    rows = grid_elem.findall("row")
                    if not rows:
                        break

                    for idx, row in enumerate(rows):
                        headers = row.findall("header")
                        cols = row.findall("col")

                        if headers:
                            if not all_headers:
                                all_headers = [
                                    h.text if h.text else "" for h in headers
                                ]
                        elif cols:
                            row_data = [c.text if c.text else "" for c in cols]
                            all_rows.append(row_data)

                    if len(all_rows) >= total_rows or len(rows) <= 1:
                        break

                    page_now += 1
                    Colors.info(
                        f"   [进度] 已获取 {len(all_rows)}/{total_rows} 行数据..."
                    )

                except ET.ParseError as e:
                    Colors.error(f"   [FAIL] XML解析错误 - {e}")
                    return (False, False)
            else:
                if "token 过期" in content or "token expired" in content.lower():
                    Colors.warning(f"   [WARN] Token已过期")
                    return (False, True)
                Colors.error(f"   [FAIL] 下载失败 - {content[:200]}")
                return (False, False)

        except requests.exceptions.RequestException as e:
            Colors.error(f"   [FAIL] 网络错误 - {e}")
            return (False, False)

    if not all_headers or not all_rows:
        Colors.error(f"   [FAIL] 未获取到数据")
        return (False, False)

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = ename[:31]

        for col_idx, header in enumerate(all_headers, 1):
            ws.cell(row=1, column=col_idx, value=header)

        for row_idx, row_data in enumerate(all_rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        output_path = os.path.join(output_dir, f"{ename}.xlsx")
        wb.save(output_path)

        file_size = os.path.getsize(output_path)
        file_size_str = (
            f"{file_size / 1024:.1f}KB"
            if file_size < 1024 * 1024
            else f"{file_size / 1024 / 1024:.1f}MB"
        )
        Colors.success(f"   [OK] {ename}.xlsx ({file_size_str}, {len(all_rows)}行)")
        return (True, False)

    except Exception as e:
        Colors.error(f"   [FAIL] 生成Excel失败 - {e}")
        return (False, False)


def run_commission_mode(args):
    if not os.path.exists(args.file):
        Colors.error(f"[FAIL] CSV文件不存在: {args.file}")
        sys.exit(1)

    if args.output != "." and not os.path.exists(args.output):
        os.makedirs(args.output, exist_ok=True)

    agents = read_csv_agents(args.file)

    if not agents:
        Colors.error("[FAIL] CSV文件中没有有效数据")
        sys.exit(1)

    total_agents = len(agents)

    Colors.info("=" * 50)
    Colors.info("BI系统线上佣金月清单下载")
    Colors.info("=" * 50)
    print()
    Colors.info(f"[CSV] CSV文件: {args.file}")
    Colors.info(f"[账期] 账期: {args.month}")
    Colors.info(f"[输出] 输出目录: {args.output}")
    Colors.info(f"[数量] 代理商数量: {total_agents}")
    print()

    username, password = load_credentials()

    if not username or not password:
        Colors.warning("[WARN] 未找到配置文件 .bi_credentials")
        Colors.info("请创建配置文件（格式如下）：")
        Colors.info("  bi_user=您的用户名")
        Colors.info("  bi_pass=您的密码")
        print()

        Colors.info("=" * 50)
        Colors.info("【安全提示】请在下方输入您的凭证")
        Colors.info("用户名和密码输入时都不会显示")
        Colors.info("=" * 50)
        print()

        username = getpass("请输入BI系统用户名: ")
        if not username:
            Colors.error("[FAIL] 用户名不能为空")
            sys.exit(1)

        password = getpass("请输入BI系统密码: ")
        if not password:
            Colors.error("[FAIL] 密码不能为空")
            sys.exit(1)
    else:
        Colors.success("[OK] 已从 .bi_credentials 加载凭证")

    print()
    Colors.info("[登录] 正在登录BI系统...")
    try:
        token = login(username, password)
        Colors.success("[OK] 登录成功")
    except Exception as e:
        Colors.error(f"[FAIL] {e}")
        sys.exit(1)

    print()
    Colors.info("=" * 50)
    Colors.info("[下载] 开始下载数据")
    Colors.info("=" * 50)
    print()

    success_count = 0
    failed_agents = []

    Colors.info("[登录] 第一轮下载...")
    token = login(username, password)
    Colors.success("[OK] 登录成功")
    time.sleep(0.5)

    for i, (agent_code, agent_name) in enumerate(agents, 1):
        Colors.info(f"[{i}/{total_agents}] {agent_name} ({agent_code})")

        success, token_expired = download_commission_data(
            token, args.month, agent_code, agent_name, args.output
        )

        if success:
            success_count += 1
        else:
            Colors.warning(f"   [WARN] 首次失败，重试中...")
            time.sleep(0.3)
            success, token_expired = download_commission_data(
                token, args.month, agent_code, agent_name, args.output
            )

            if success:
                success_count += 1
            else:
                failed_agents.append((agent_code, agent_name))

    max_retry_rounds = 3
    retry_round = 0

    while failed_agents and retry_round < max_retry_rounds:
        retry_round += 1
        print()
        Colors.warning(
            f"[重试] 第 {retry_round} 轮重试，剩余 {len(failed_agents)} 个失败..."
        )

        try:
            token = login(username, password)
            Colors.success("[OK] 重新登录成功")
            time.sleep(0.5)
        except Exception as e:
            Colors.error(f"[FAIL] 登录失败: {e}")
            break

        still_failed = []
        for i, (agent_code, agent_name) in enumerate(failed_agents, 1):
            Colors.info(f"[重试 {i}/{len(failed_agents)}] {agent_name} ({agent_code})")

            success, token_expired = download_commission_data(
                token, args.month, agent_code, agent_name, args.output
            )

            if success:
                success_count += 1
            else:
                time.sleep(0.3)
                success, token_expired = download_commission_data(
                    token, args.month, agent_code, agent_name, args.output
                )
                if success:
                    success_count += 1
                else:
                    still_failed.append((agent_code, agent_name))

        failed_agents = still_failed

    print()
    Colors.success("=" * 50)
    Colors.success("[完成] 下载完成！")
    Colors.success("=" * 50)
    print()
    Colors.info("[统计]:")
    Colors.success(f"   [OK] 成功: {success_count}/{total_agents}")
    if failed_agents:
        Colors.error(f"   [FAIL] 失败: {len(failed_agents)}")
        for code, name in failed_agents:
            Colors.error(f"      - {name} ({code})")
    Colors.info(f"   [输出] 输出目录: {args.output}")


def run_daily_mode(args):
    if args.output != "." and not os.path.exists(args.output):
        os.makedirs(args.output, exist_ok=True)

    Colors.info("=" * 50)
    Colors.info("BI系统电渠监控日表下载")
    Colors.info("=" * 50)
    print()
    Colors.info(f"[账期] 日账期: {args.date}")
    Colors.info(f"[输出] 输出目录: {args.output}")
    print()

    username, password = load_credentials()

    if not username or not password:
        Colors.warning("[WARN] 未找到配置文件 .bi_credentials")
        Colors.info("请创建配置文件（格式如下）：")
        Colors.info("  bi_user=您的用户名")
        Colors.info("  bi_pass=您的密码")
        print()
        Colors.info("=" * 50)
        Colors.info("【安全提示】请在下方输入您的凭证")
        Colors.info("用户名和密码输入时都不会显示")
        Colors.info("=" * 50)
        print()

        username = getpass("请输入BI系统用户名: ")
        if not username:
            Colors.error("[FAIL] 用户名不能为空")
            sys.exit(1)

        password = getpass("请输入BI系统密码: ")
        if not password:
            Colors.error("[FAIL] 密码不能为空")
            sys.exit(1)
    else:
        Colors.success("[OK] 已从 .bi_credentials 加载凭证")

    print()
    Colors.info("[登录] 正在登录BI系统...")
    try:
        token = login(username, password)
        Colors.success("[OK] 登录成功")
    except Exception as e:
        Colors.error(f"[FAIL] {e}")
        sys.exit(1)

    print()
    Colors.info("=" * 50)
    Colors.info("[下载] 开始下载数据")
    Colors.info("=" * 50)
    print()

    time.sleep(0.5)

    success, token_expired = download_daily_data(token, args.date, args.output)

    if not success:
        Colors.warning("[WARN] 首次失败，重试中...")
        time.sleep(0.5)

        try:
            token = login(username, password)
            Colors.success("[OK] 重新登录成功")
            time.sleep(0.5)
        except Exception as e:
            Colors.error(f"[FAIL] 登录失败: {e}")
            sys.exit(1)

        success, token_expired = download_daily_data(token, args.date, args.output)

    print()
    if success:
        Colors.success("=" * 50)
        Colors.success("[完成] 下载完成！")
        Colors.success("=" * 50)
        Colors.info(f"   [输出] 输出目录: {args.output}")
    else:
        Colors.error("=" * 50)
        Colors.error("[FAIL] 下载失败")
        Colors.error("=" * 50)


def run_sales_monthly_mode(args):
    if args.output != "." and not os.path.exists(args.output):
        os.makedirs(args.output, exist_ok=True)

    Colors.info("=" * 50)
    Colors.info("BI系统销售点月报下载")
    Colors.info("=" * 50)
    print()
    Colors.info(f"[账期] 月账期: {args.month}")
    Colors.info(f"[输出] 输出目录: {args.output}")
    print()

    username, password = load_credentials()

    if not username or not password:
        Colors.warning("[WARN] 未找到配置文件 .bi_credentials")
        Colors.info("请创建配置文件（格式如下）：")
        Colors.info("  bi_user=您的用户名")
        Colors.info("  bi_pass=您的密码")
        print()
        Colors.info("=" * 50)
        Colors.info("【安全提示】请在下方输入您的凭证")
        Colors.info("用户名和密码输入时都不会显示")
        Colors.info("=" * 50)
        print()

        username = getpass("请输入BI系统用户名: ")
        if not username:
            Colors.error("[FAIL] 用户名不能为空")
            sys.exit(1)

        password = getpass("请输入BI系统密码: ")
        if not password:
            Colors.error("[FAIL] 密码不能为空")
            sys.exit(1)
    else:
        Colors.success("[OK] 已从 .bi_credentials 加载凭证")

    print()
    Colors.info("[登录] 正在登录BI系统...")
    try:
        token = login(username, password)
        Colors.success("[OK] 登录成功")
    except Exception as e:
        Colors.error(f"[FAIL] {e}")
        sys.exit(1)

    print()
    Colors.info("=" * 50)
    Colors.info("[下载] 开始下载数据")
    Colors.info("=" * 50)
    print()

    time.sleep(0.5)

    success, token_expired = download_sales_monthly_data(token, args.month, args.output)

    if not success:
        Colors.warning("[WARN] 首次失败，重试中...")
        time.sleep(0.5)

        try:
            token = login(username, password)
            Colors.success("[OK] 重新登录成功")
            time.sleep(0.5)
        except Exception as e:
            Colors.error(f"[FAIL] 登录失败: {e}")
            sys.exit(1)

        success, token_expired = download_sales_monthly_data(
            token, args.month, args.output
        )

    print()
    if success:
        Colors.success("=" * 50)
        Colors.success("[完成] 下载完成！")
        Colors.success("=" * 50)
        Colors.info(f"   [输出] 输出目录: {args.output}")
    else:
        Colors.error("=" * 50)
        Colors.error("[FAIL] 下载失败")
        Colors.error("=" * 50)


def run_hr_mode(args):
    if args.output != "." and not os.path.exists(args.output):
        os.makedirs(args.output, exist_ok=True)

    dataset_info = HR_DATASETS.get(args.name)
    if not dataset_info:
        Colors.error(f"[FAIL] 未知的数据集名称: {args.name}")
        Colors.info("可用的数据集:")
        for key, info in HR_DATASETS.items():
            Colors.info(f"  - {key}: {info['name']}")
        sys.exit(1)

    qry_type = args.qry_type if args.qry_type else "sql"

    Colors.info("=" * 50)
    Colors.info("BI系统HR自服务数据集下载")
    Colors.info("=" * 50)
    print()
    Colors.info(f"[数据集] {dataset_info['name']}")
    Colors.info(f"[路径] {dataset_info['path']}")
    Colors.info(f"[类型] {qry_type}")
    Colors.info(f"[输出] 输出目录: {args.output}")
    print()

    username, password = load_credentials()

    if not username or not password:
        Colors.warning("[WARN] 未找到配置文件 .bi_credentials")
        Colors.info("请创建配置文件（格式如下）：")
        Colors.info("  bi_user=您的用户名")
        Colors.info("  bi_pass=您的密码")
        print()
        Colors.info("=" * 50)
        Colors.info("【安全提示】请在下方输入您的凭证")
        Colors.info("用户名和密码输入时都不会显示")
        Colors.info("=" * 50)
        print()

        username = getpass("请输入BI系统用户名: ")
        if not username:
            Colors.error("[FAIL] 用户名不能为空")
            sys.exit(1)

        password = getpass("请输入BI系统密码: ")
        if not password:
            Colors.error("[FAIL] 密码不能为空")
            sys.exit(1)
    else:
        Colors.success("[OK] 已从 .bi_credentials 加载凭证")

    print()
    Colors.info("[登录] 正在登录BI系统...")
    try:
        token = login(username, password)
        Colors.success("[OK] 登录成功")
    except Exception as e:
        Colors.error(f"[FAIL] {e}")
        sys.exit(1)

    print()
    Colors.info("=" * 50)
    Colors.info("[下载] 开始下载数据")
    Colors.info("=" * 50)
    print()

    time.sleep(0.5)

    success, token_expired = download_dataflow_data(
        token, dataset_info["path"], dataset_info["name"], args.output, qry_type
    )

    if not success:
        Colors.warning("[WARN] 首次失败，重试中...")
        time.sleep(0.5)

        try:
            token = login(username, password)
            Colors.success("[OK] 重新登录成功")
            time.sleep(0.5)
        except Exception as e:
            Colors.error(f"[FAIL] 登录失败: {e}")
            sys.exit(1)

        success, token_expired = download_dataflow_data(
            token, dataset_info["path"], dataset_info["name"], args.output, qry_type
        )

    print()
    if success:
        Colors.success("=" * 50)
        Colors.success("[完成] 下载完成！")
        Colors.success("=" * 50)
        Colors.info(f"   [输出] 输出目录: {args.output}")
    else:
        Colors.error("=" * 50)
        Colors.error("[FAIL] 下载失败")
        Colors.error("=" * 50)


def run_report_mode(args):
    if args.output != "." and not os.path.exists(args.output):
        os.makedirs(args.output, exist_ok=True)

    Colors.info("=" * 50)
    Colors.info("BI系统数据报告下载")
    Colors.info("=" * 50)
    print()
    Colors.info(f"[报告路径] {args.path}")
    Colors.info(f"[图表名称] {args.ename}")
    Colors.info(f"[输出] 输出目录: {args.output}")
    print()

    username, password = load_credentials()

    if not username or not password:
        Colors.warning("[WARN] 未找到配置文件 .bi_credentials")
        Colors.info("请创建配置文件（格式如下）：")
        Colors.info("  bi_user=您的用户名")
        Colors.info("  bi_pass=您的密码")
        print()
        Colors.info("=" * 50)
        Colors.info("【安全提示】请在下方输入您的凭证")
        Colors.info("用户名和密码输入时都不会显示")
        Colors.info("=" * 50)
        print()

        username = getpass("请输入BI系统用户名: ")
        if not username:
            Colors.error("[FAIL] 用户名不能为空")
            sys.exit(1)

        password = getpass("请输入BI系统密码: ")
        if not password:
            Colors.error("[FAIL] 密码不能为空")
            sys.exit(1)
    else:
        Colors.success("[OK] 已从 .bi_credentials 加载凭证")

    print()
    Colors.info("[登录] 正在登录BI系统...")
    try:
        token = login(username, password, REPORT_API_BASE)
        Colors.success("[OK] 登录成功")
    except Exception as e:
        Colors.error(f"[FAIL] {e}")
        sys.exit(1)

    print()
    Colors.info("=" * 50)
    Colors.info("[下载] 开始下载数据")
    Colors.info("=" * 50)
    print()

    time.sleep(0.5)

    success, token_expired = download_report_data(
        token, args.path, args.ename, args.output
    )

    if not success:
        Colors.warning("[WARN] 首次失败，重试中...")
        time.sleep(0.5)

        try:
            token = login(username, password, REPORT_API_BASE)
            Colors.success("[OK] 重新登录成功")
            time.sleep(0.5)
        except Exception as e:
            Colors.error(f"[FAIL] 登录失败: {e}")
            sys.exit(1)

        success, token_expired = download_report_data(
            token, args.path, args.ename, args.output
        )

    print()
    if success:
        Colors.success("=" * 50)
        Colors.success("[完成] 下载完成！")
        Colors.success("=" * 50)
        Colors.info(f"   [输出] 输出目录: {args.output}")
    else:
        Colors.error("=" * 50)
        Colors.error("[FAIL] 下载失败")
        Colors.error("=" * 50)


def is_token_expired(t: str) -> bool:
    t = t.lower()
    return "token 过期" in t or "token expired" in t


def is_xml_error(t: str) -> bool:
    if "丢弃空的参数失败" in t:
        return True
    t = t.strip()
    if not (t.startswith("<?xml") or t.startswith("<results")):
        return False
    try:
        root = ET.fromstring(t)
        for r in root.findall(".//result"):
            msg = r.find("message")
            lvl = r.find("level")
            if msg is not None and msg.text and msg.text.lower() not in ("null", ""):
                return True
            if lvl is not None and lvl.text:
                try:
                    level_val = int(lvl.text)
                    if level_val >= 6:
                        return True
                except:
                    pass
        return False
    except:
        return True


def parse_xml_error(t: str) -> dict:
    try:
        root = ET.fromstring(t)
        for r in root.findall(".//result"):
            msg = r.find("message")
            if msg is None:
                msg = r.find("level")
            err_msg = msg.text if msg is not None else t[:200]
            return {"errors": [{"message": err_msg}]}
    except ET.ParseError:
        pass
    return {"errors": [{"message": t[:200]}]}


def parse_missing_params_from_error(error_msg: str) -> list:
    if "丢弃空的参数失败" in error_msg:
        m = re.search(
            r'[\u201c\u201d""\'"]([\u4e00-\u9fa5]+)[\u201d\u201c""\'"]', error_msg
        )
        if m:
            return [m.group(1)]
        m2 = re.search(r"：\s*([\u4e00-\u9fa5]+)", error_msg)
        if m2:
            return [m2.group(1)]
    patterns = [
        r"缺少[参数：:](.+)",
        r"请输入[参数\[](.+?)[\]]",
        r"缺少\s+(\w+)\s+参数",
        r"需要\s+(\w+)\s+参数",
    ]
    params = []
    for pattern in patterns:
        matches = re.findall(pattern, error_msg)
        params.extend(matches)
    return [p.strip() for p in params if p.strip()]


def format_size(s: int) -> str:
    if s < 1048576:
        return f"{s / 1024:.1f}KB"
    return f"{s / 1024 / 1024:.1f}MB"


def is_excel(b: bytes) -> bool:
    return len(b) > 4 and (b[:2] == b"PK" or b[:4] == b"\xd0\xcf\x11\xe0")


def is_xlsx(b: bytes) -> bool:
    return len(b) > 4 and b[:2] == b"PK"


def response_preview(content: bytes) -> str:
    text = content[:500].decode("utf-8", errors="replace").strip()
    if text:
        return text
    return content[:80].hex()


def write_download_response(output_file: str, content: bytes, file_type: str) -> None:
    expected_type = str(file_type or "").lower()
    if expected_type == "xlsx" and not is_xlsx(content):
        preview = response_preview(content)
        raise ValueError(
            "BI 返回内容不是标准 .xlsx，已停止保存。请确认 BI 下载格式固定为 xlsx。"
            f"返回内容预览: {preview[:500]}"
        )
    with open(output_file, "wb") as f:
        f.write(content)


def clean_excel_header(value) -> str:
    return str(value or "").replace("\r", "").replace("\n", "").replace("\t", "").strip()


def normalize_area_name(value) -> str:
    text = str(value or "").strip()
    return text[:-1] if text.endswith("市") else text


def read_xlsx_shape_and_rows(path: str) -> tuple[list[str], list[tuple]]:
    handle = open(path, "rb")
    try:
        wb = openpyxl.load_workbook(handle, read_only=True, data_only=True)
        try:
            if not wb.sheetnames:
                raise ValueError("Excel 文件没有 sheet")
            ws = wb[wb.sheetnames[0]]
            iterator = ws.iter_rows(values_only=True)
            try:
                raw_headers = next(iterator)
            except StopIteration as exc:
                raise ValueError("Excel 文件为空：没有表头") from exc
            headers = [clean_excel_header(value) for value in raw_headers]
            if not any(headers):
                raise ValueError("Excel 第 1 行表头为空")
            rows = [row for row in iterator if any(value not in (None, "") for value in row)]
            return headers, rows
        finally:
            wb.close()
    finally:
        handle.close()


def validate_saved_download(config_name: str, output_file: str, file_type: str) -> str:
    if str(file_type or "").lower() != "xlsx":
        return "非 xlsx 下载，已跳过 Excel 内容校验"
    if not os.path.exists(output_file):
        raise ValueError(f"下载文件未保存成功: {output_file}")
    if os.path.getsize(output_file) <= 0:
        raise ValueError(f"下载文件为空: {output_file}")

    headers, rows = read_xlsx_shape_and_rows(output_file)
    indexes = {header: idx for idx, header in enumerate(headers) if header}

    if config_name == "field-service-agent-area-summary":
        if "地市" not in indexes:
            raise ValueError(f"地市汇总文件缺少字段: 地市；实际字段: {'、'.join(headers)}")
        city_idx = indexes["地市"]
        cities = [
            normalize_area_name(row[city_idx] if city_idx < len(row) else None)
            for row in rows
        ]
        detail_cities = sorted({city for city in cities if city and city != "全省"})
        if not detail_cities:
            raise ValueError("地市汇总文件没有具体地市数据，不能只有全省汇总")
        return f"地市汇总校验通过：{len(rows)} 行，具体地市 {len(detail_cities)} 个"

    if config_name == "field-service-agent-city-summary":
        missing = [field for field in ("地市", "区县") if field not in indexes]
        if missing:
            raise ValueError(f"区县文件缺少字段: {', '.join(missing)}；实际字段: {'、'.join(headers)}")
        county_idx = indexes["区县"]
        county_rows = [
            row for row in rows
            if county_idx < len(row) and str(row[county_idx] or "").strip()
        ]
        if not county_rows:
            raise ValueError("区县文件没有有效数据行：区县列为空或只有表头")
        return f"区县文件校验通过：{len(rows)} 行，有效区县行 {len(county_rows)} 行"

    return f"Excel 保存校验通过：{len(rows)} 行"


def output_pattern_for_existing(template: str, ext: str, params: dict) -> re.Pattern:
    tokens = build_output_tokens(params)
    tokens["ext"] = ext
    pattern = re.escape(template)
    replacements = {
        "date": rf"{re.escape(tokens['acct_day'])}_{re.escape(tokens['month_id'])}_\d{{14}}",
        "acct_day": re.escape(tokens["acct_day"]),
        "month_id": re.escape(tokens["month_id"]),
        "timestamp": r"\d{14}",
        "ext": re.escape(ext),
    }
    for key, value in replacements.items():
        pattern = pattern.replace(re.escape(f"{{{key}}}"), value)
    return re.compile(rf"^{pattern}$")


def find_valid_existing_download(config_name: str, output_dir: str, template: str, ext: str, params: dict) -> str | None:
    if not config_name.startswith("field-service-agent-"):
        return None
    if str(ext or "").lower() != "xlsx" or not os.path.isdir(output_dir):
        return None

    pattern = output_pattern_for_existing(template, ext, params)
    candidates = []
    for name in os.listdir(output_dir):
        if name.startswith("~$") or not pattern.match(name):
            continue
        path = os.path.join(output_dir, name)
        if os.path.isfile(path):
            candidates.append(path)

    for path in sorted(candidates, key=os.path.getmtime, reverse=True):
        try:
            validate_saved_download(config_name, path, ext)
            return path
        except Exception as exc:
            Colors.warning(f"[跳过] 本地文件未通过保存校验: {path}；{exc}")
    return None


def parse_params_string(params_str: str) -> dict:
    if not params_str:
        return {}
    params = {}
    for item in params_str.split(","):
        item = item.strip()
        if not item:
            continue
        if "=" in item:
            key, val = item.split("=", 1)
            key = key.strip()
            val = val.strip()
            if ":" in val:
                val, typ = val.split(":", 1)
                params[key] = (val, typ)
            elif key.endswith("_id") or key.endswith("_type") or key.endswith("_flag"):
                params[key] = (val, "string")
            elif val.isdigit():
                params[key] = (val, "string")
            else:
                params[key] = (val, "string")
    return params


def get_skill_dir() -> str:
    return os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def get_default_config_file() -> str:
    return os.path.join(get_skill_dir(), "download_configs.json")


def load_configs(config_file: str = None) -> dict:
    if config_file is None:
        config_file = get_default_config_file()
    if not os.path.exists(config_file):
        return {}
    with open(config_file, "r", encoding="utf-8") as f:
        return json.load(f).get("configs", {})


def save_discovered_params(
    config_name: str,
    path: str,
    params: list,
    item_type: str,
    config_file: str = None,
):
    if config_file is None:
        config_file = get_default_config_file()
    full_config = {}
    if os.path.exists(config_file):
        with open(config_file, "r", encoding="utf-8") as f:
            full_config = json.load(f)

    if "configs" not in full_config:
        full_config["configs"] = {}

    param_defs = {}
    for p in params:
        param_defs[p] = ["string", ""]

    full_config["configs"][config_name] = {
        "type": item_type,
        "path": path,
        "params": param_defs,
    }

    with open(config_file, "w", encoding="utf-8") as f:
        json.dump(full_config, f, ensure_ascii=False, indent=2)

    Colors.success(f"已保存配置到 {config_file}")


def discover_params(
    path: str,
    item_type: str,
    api_base: str = API_BASE,
    qry_type: str = "sql",
    file_type: str = "xlsx",
    report_type: str = "db",
) -> list:
    username, password = load_credentials()
    if not username or not password:
        Colors.error("[FAIL] 未找到凭证")
        return []

    token = login(username, password, api_base)

    params_xml = f"""<?xml version="1.0" encoding="UTF-8"?>
<info>
    <qryPath>{path}</qryPath>
    <qryType>{qry_type}</qryType>
    <fileType>{file_type}</fileType>
</info>"""

    url = f"{api_base}?action=downloadQuery&token={token}"

    try:
        resp = SESSION.post(
            url, data={"xmlData": params_xml}, verify=False, timeout=120
        )

        if is_token_expired(resp.text):
            token = login(username, password, api_base)
            resp = SESSION.post(
                url, data={"xmlData": params_xml}, verify=False, timeout=120
            )

        if is_xml_error(resp.text):
            error_msg = resp.text
            missing_params = parse_missing_params_from_error(error_msg)
            return missing_params

        return []

    except Exception as e:
        Colors.error(f"[FAIL] 参数发现失败: {e}")
        return []


def run_discover_mode(args):
    if not args.path:
        Colors.error("[FAIL] discover模式需要 --path 参数")
        sys.exit(1)

    item_type = "report" if args.report_type else "query"
    api_base = REPORT_API_BASE if item_type == "report" else API_BASE

    Colors.info("=" * 50)
    Colors.info("BI系统参数发现模式")
    Colors.info("=" * 50)
    print()
    Colors.info(f"[路径] {args.path}")
    Colors.info(f"[类型] {item_type}")
    print()

    missing = discover_params(
        args.path,
        item_type,
        api_base=api_base,
        qry_type=args.qry_type,
        file_type=args.file_type,
        report_type=args.report_type or "db",
    )

    if not missing:
        Colors.success("未发现需要参数，可直接下载")

        username, password = load_credentials()
        token = login(username, password, api_base)

        if item_type == "query":
            xml_data = f"""<?xml version="1.0" encoding="UTF-8"?>
<info>
    <qryPath>{args.path}</qryPath>
    <qryType>{args.qry_type}</qryType>
    <fileType>{args.file_type}</fileType>
</info>"""
        else:
            xml_data = f"""<?xml version="1.0" encoding="UTF-8"?>
<info>
    <dbpath>{args.path}</dbpath>
    <type>{args.report_type or "db"}</type>
    <fileType>{args.file_type}</fileType>
</info>"""

        url = f"{api_base}?action=downloadQuery&token={token}"
        resp = SESSION.post(url, data={"xmlData": xml_data}, verify=False, timeout=120)

        if is_xml_error(resp.text):
            Colors.error(f"[FAIL] 下载失败: {resp.text[:500]}")
            return

        output_file = os.path.join(args.output, f"output.{args.file_type}")
        write_download_response(output_file, resp.content, args.file_type)
        Colors.success(f"[OK] -> {output_file} ({format_size(len(resp.content))})")
        return

    Colors.warning(f"发现需要以下参数: {missing}")

    username, password = load_credentials()
    token = login(username, password, api_base)

    today = datetime.today()
    last_month = today - relativedelta(months=1)
    yesterday = today - timedelta(days=1)

    auto_values = {
        "last_month": last_month.strftime("%Y%m"),
        "current_month": today.strftime("%Y%m"),
        "yesterday": yesterday.strftime("%Y%m%d"),
        "today": today.strftime("%Y%m%d"),
    }

    def get_candidates_for_param(param_name):
        pn = param_name.lower()
        if any(k in pn for k in ["账期"]):
            return [
                auto_values["yesterday"],
                auto_values["last_month"],
                auto_values["current_month"],
            ]
        elif any(k in pn for k in ["日", "天", "date", "day"]):
            return [auto_values["yesterday"], auto_values["today"]]
        elif any(k in pn for k in ["月", "月份"]):
            return [auto_values["last_month"], auto_values["current_month"]]
        elif any(k in pn for k in ["年", "year"]):
            return [str(today.year)]
        return []

    param_candidates = {}
    auto_params = {}
    for p in missing:
        candidates = get_candidates_for_param(p)
        if candidates:
            param_candidates[p] = candidates
            auto_params[p] = candidates[0]
            Colors.info(
                f"  [自动] {p} -> 尝试 {[candidates[0], candidates[1] if len(candidates) > 1 else '...']}"
            )
        else:
            user_input = input(f"请输入 {p}: ").strip()
            if user_input:
                auto_params[p] = user_input
            else:
                Colors.warning(f"  跳过参数: {p}")

    if not auto_params:
        Colors.error("[FAIL] 没有提供任何参数")
        return

    Colors.info("\n自动尝试下载...")

    def build_xml(params):
        p_str = "".join(
            f"<param><name>{k}</name><type>string</type><value>{v}</value></param>"
            for k, v in params.items()
        )
        if item_type == "query":
            return f"""<?xml version="1.0" encoding="UTF-8"?>
<info>
    <qryPath>{args.path}</qryPath>
    <qryType>{args.qry_type}</qryType>
    <fileType>{args.file_type}</fileType>
    <params>{p_str}</params>
</info>"""
        else:
            return f"""<?xml version="1.0" encoding="UTF-8"?>
<info>
    <dbpath>{args.path}</dbpath>
    <type>{args.report_type or "db"}</type>
    <fileType>{args.file_type}</fileType>
    <params>{p_str}</params>
</info>"""

    url = f"{api_base}?action=downloadQuery&token={token}"
    max_attempts = 3
    attempt = 0

    while attempt < max_attempts:
        attempt += 1
        xml_data = build_xml(auto_params)

        try:
            resp = SESSION.post(
                url, data={"xmlData": xml_data}, verify=False, timeout=120
            )

            if is_xml_error(resp.text):
                if attempt >= max_attempts:
                    Colors.error(
                        f"[FAIL] 下载失败 (已尝试{attempt}次): {resp.text[:300]}"
                    )
                    return
                Colors.warning(f"  第{attempt}次尝试失败，尝试其他参数组合...")
                changed = False
                for p, candidates in param_candidates.items():
                    idx = (
                        candidates.index(auto_params[p])
                        if auto_params[p] in candidates
                        else 0
                    )
                    if idx + 1 < len(candidates):
                        auto_params[p] = candidates[idx + 1]
                        Colors.info(f"    {p} -> {auto_params[p]}")
                        changed = True
                if not changed:
                    Colors.error(f"[FAIL] 所有参数组合已尝试")
                    return
                continue

            output_file = os.path.join(args.output, f"output.{args.file_type}")
            write_download_response(output_file, resp.content, args.file_type)
            Colors.success(f"[OK] -> {output_file} ({format_size(len(resp.content))})")

            if args.save:
                save_discovered_params(
                    args.name or "discovered", args.path, missing, item_type
                )
            return

        except Exception as e:
            Colors.error(f"[FAIL] {e}")
            return

    if args.save:
        save_discovered_params(args.name or "discovered", args.path, missing, item_type)


def run_config_mode(args):
    configs = load_configs()
    config = configs.get(args.name)

    if not config:
        Colors.error(f"[FAIL] 未找到配置: {args.name}")
        Colors.info("可用配置: " + ", ".join(configs.keys()))
        sys.exit(1)

    params = parse_params_string(args.params)
    if args.name.startswith("key-product-report-") and "acc_day" not in params:
        params["acc_day"] = (get_yesterday(), "string")
    try:
        validate_field_service_params(args.name, params)
    except ValueError as e:
        Colors.error(f"[FAIL] {e}")
        sys.exit(1)

    output_template = config.get("output_template", "output.{ext}")
    ext = config.get("file_type", "xlsx")
    existing_file = None if args.force else find_valid_existing_download(
        args.name, args.output, output_template, ext, params
    )
    if existing_file:
        Colors.success(f"[复用] 本地文件已通过保存校验 -> {existing_file}")
        return

    Colors.info("=" * 50)
    Colors.info(f"配置模式: {args.name}")
    Colors.info("=" * 50)
    print()
    Colors.info(f"[路径] {config['path']}")
    Colors.info(f"[参数] {params}")
    print()

    username, password = load_credentials()
    if not username or not password:
        Colors.error("[FAIL] 未找到凭证")
        sys.exit(1)

    token = login(username, password)

    item_type = "report" if config.get("type") == "report" else "query"
    api_base = REPORT_API_BASE if item_type == "report" else API_BASE

    p_str = ""
    for k, v in params.items():
        if isinstance(v, tuple):
            val, typ = v
        else:
            val, typ = v, "string"
        p_str += (
            f"<param><name>{k}</name><type>{typ}</type><value>{val}</value></param>"
        )

    if item_type == "query":
        xml_data = f"""<?xml version="1.0" encoding="UTF-8"?>
<info>
    <qryPath>{config["path"]}</qryPath>
    <qryType>{config.get("qry_type", "sql")}</qryType>
    <fileType>{config.get("file_type", "xlsx")}</fileType>
    <params>{p_str}</params>
</info>"""
        url = f"{api_base}?action=downloadQuery&token={token}"
    elif "ename" in config:
        # 图表下载模式（使用 getElemData）
        params_xml = ""
        if config.get("params"):
            for k, v in config["params"].items():
                val = params.get(k, v[1]) if isinstance(v, tuple) else params.get(k, v)
                params_xml += f"<param><name>{k}</name><type>string</type><value>{val}</value></param>"
        else:
            # 自动添加账期参数
            acct_month = params.get(
                "月账期", params.get("账期", args.month if args.month else "202601")
            )
            params_xml = f"<param><name>账期</name><type>string</type><value>{acct_month}</value></param>"

        xml_data = f"""<?xml version="1.0" encoding="UTF-8"?>
<info>
    <page>
        <pageSize>10000</pageSize>
        <pageNow>1</pageNow>
    </page>
    <getdata>
        <dbpath>{config["path"]}</dbpath>
        <ename>{config["ename"]}</ename>
        {params_xml}
    </getdata>
</info>"""
        url = f"{REPORT_API_BASE}?action=getElemData&token={token}"
    else:
        # 传统文件下载模式
        xml_data = f"""<?xml version="1.0" encoding="UTF-8"?>
<info>
    <dbpath>{config["path"]}</dbpath>
    <type>{config.get("report_type", "db")}</type>
    <fileType>{config.get("file_type", "pdf")}</fileType>
    <params>{p_str}</params>
</info>"""
        url = f"{api_base}?action=downloadFile&token={token}"

    try:
        resp = SESSION.post(url, data={"xmlData": xml_data}, verify=False, timeout=120)

        if is_token_expired(resp.text):
            token = login(username, password, api_base)
            resp = SESSION.post(
                url, data={"xmlData": xml_data}, verify=False, timeout=120
            )

        if is_xml_error(resp.text):
            error_msg = resp.text
            missing = parse_missing_params_from_error(error_msg)
            if missing:
                Colors.warning(f"缺少参数: {missing}")
                for p in missing:
                    if p not in params:
                        params[p] = (input(f"请输入 {p}: ").strip(), "string")

                p_str = ""
                for k, v in params.items():
                    if isinstance(v, tuple):
                        val, typ = v
                    else:
                        val, typ = v, "string"
                    p_str += f"<param><name>{k}</name><type>{typ}</type><value>{val}</value></param>"

                if item_type == "query":
                    xml_data = f"""<?xml version="1.0" encoding="UTF-8"?>
<info>
    <qryPath>{config["path"]}</qryPath>
    <qryType>{config.get("qry_type", "sql")}</qryType>
    <fileType>{config.get("file_type", "xlsx")}</fileType>
    <params>{p_str}</params>
</info>"""
                else:
                    xml_data = f"""<?xml version="1.0" encoding="UTF-8"?>
<info>
    <dbpath>{config["path"]}</dbpath>
    <type>{config.get("report_type", "db")}</type>
    <fileType>{config.get("file_type", "pdf")}</fileType>
    <params>{p_str}</params>
</info>"""
                resp = SESSION.post(
                    url, data={"xmlData": xml_data}, verify=False, timeout=120
                )

        if is_xml_error(resp.text):
            Colors.error(f"[FAIL] {resp.text[:500]}")
            return

        output_file = os.path.join(
            args.output,
            render_output_template(output_template, ext, params),
        )
        temp_output_file = f"{output_file}.part"
        invalid_output_file = f"{output_file}.invalid"
        write_download_response(temp_output_file, resp.content, ext)
        try:
            validation_message = validate_saved_download(args.name, temp_output_file, ext)
        except Exception:
            if os.path.exists(temp_output_file):
                os.replace(temp_output_file, invalid_output_file)
            raise
        os.replace(temp_output_file, output_file)
        Colors.success(f"[OK] -> {output_file} ({format_size(len(resp.content))})")
        Colors.success(f"[校验] {validation_message}")

    except Exception as e:
        Colors.error(f"[FAIL] {e}")


def run_direct_mode(args):
    username, password = load_credentials()
    if not username or not password:
        Colors.error("[FAIL] 未找到凭证")
        sys.exit(1)

    params = parse_params_string(args.params)

    Colors.info("=" * 50)
    Colors.info("BI系统直接下载模式")
    Colors.info("=" * 50)
    print()
    Colors.info(f"[路径] {args.path}")
    Colors.info(f"[参数] {params}")
    print()

    token = login(username, password, REPORT_API_BASE if args.report_type else API_BASE)

    p_str = ""
    for k, v in params.items():
        if isinstance(v, tuple):
            val, typ = v
        else:
            val, typ = v, "string"
        p_str += (
            f"<param><name>{k}</name><type>{typ}</type><value>{val}</value></param>"
        )

    if args.report_type:
        xml_data = f"""<?xml version="1.0" encoding="UTF-8"?>
<info>
    <dbpath>{args.path}</dbpath>
    <type>{args.report_type}</type>
    <fileType>{args.file_type}</fileType>
    <params>{p_str}</params>
</info>"""
        url = f"{REPORT_API_BASE}?action=downloadFile&token={token}"
    else:
        xml_data = f"""<?xml version="1.0" encoding="UTF-8"?>
<info>
    <qryPath>{args.path}</qryPath>
    <qryType>{args.qry_type}</qryType>
    <fileType>{args.file_type}</fileType>
    <params>{p_str}</params>
</info>"""
        url = f"{API_BASE}?action=downloadQuery&token={token}"

    try:
        resp = SESSION.post(url, data={"xmlData": xml_data}, verify=False, timeout=120)

        if is_token_expired(resp.text):
            token = login(
                username, password, REPORT_API_BASE if args.report_type else API_BASE
            )
            resp = SESSION.post(
                url, data={"xmlData": xml_data}, verify=False, timeout=120
            )

        if is_xml_error(resp.text):
            error_msg = resp.text
            missing = parse_missing_params_from_error(error_msg)
            if missing:
                Colors.warning(f"缺少参数: {missing}")
                for p in missing:
                    if p not in params:
                        params[p] = (input(f"请输入 {p}: ").strip(), "string")

                p_str = ""
                for k, v in params.items():
                    if isinstance(v, tuple):
                        val, typ = v
                    else:
                        val, typ = v, "string"
                    p_str += f"<param><name>{k}</name><type>{typ}</type><value>{val}</value></param>"

                if args.report_type:
                    xml_data = f"""<?xml version="1.0" encoding="UTF-8"?>
<info>
    <dbpath>{args.path}</dbpath>
    <type>{args.report_type}</type>
    <fileType>{args.file_type}</fileType>
    <params>{p_str}</params>
</info>"""
                else:
                    xml_data = f"""<?xml version="1.0" encoding="UTF-8"?>
<info>
    <qryPath>{args.path}</qryPath>
    <qryType>{args.qry_type}</qryType>
    <fileType>{args.file_type}</fileType>
    <params>{p_str}</params>
</info>"""
                resp = SESSION.post(
                    url, data={"xmlData": xml_data}, verify=False, timeout=120
                )

        if is_xml_error(resp.text):
            Colors.error(f"[FAIL] {resp.text[:500]}")
            return

        output_file = os.path.join(args.output, f"output.{args.file_type}")
        write_download_response(output_file, resp.content, args.file_type)
        Colors.success(f"[OK] -> {output_file} ({format_size(len(resp.content))})")

    except Exception as e:
        Colors.error(f"[FAIL] {e}")


def main():
    parser = argparse.ArgumentParser(
        description="BI系统数据下载工具",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
下载类型:
  commission      - 线上佣金出账月清单
  daily          - 电渠监控日表
  hr             - HR自服务数据集
  report         - 数据报告
  sales_monthly  - 销售点月报
  discover       - 智能参数发现（推荐新数据集）
  config         - 使用预配置下载
  direct         - 直接指定路径和参数下载

HR数据集名称 (仅hr模式):
  channel_staff      - 渠道三大岗位人员清单
  manager_eval       - 渠道经理人员评价表
  direct_sales_eval  - 公众渠道直销经理人员评价表
  sales_rep_eval     - 营业代表人员评价表

CSV文件格式 (仅佣金月清单模式):
  ORG_ID,OPERATORS_NBR,OPERATORS_NAME,批次

示例:
  # 下载HR人效数据
  python download_bi_data.py -t report -p "省公司报表/人力资源部/人员效能分析/渠道相关" -e "图表1" -o ./downloads

  # 下载店效销售点月报
  python download_bi_data.py -t sales_monthly -m 202601 -o ./downloads

  # 发现新数据集参数
  python download_bi_data.py -t discover --path "省公司数据集/xxx" -o ./downloads

  # 直接下载
  python download_bi_data.py -t direct --path "省公司数据集/xxx" --qry-type sql -p "参数1=value1" -o ./downloads
        """,
    )

    parser.add_argument(
        "-t",
        "--type",
        default="commission",
        choices=[
            "commission",
            "daily",
            "hr",
            "report",
            "sales_monthly",
            "discover",
            "config",
            "direct",
        ],
        help="下载类型",
    )
    parser.add_argument(
        "-f", "--file", default="dls.csv", help="CSV文件路径 (仅佣金月清单模式)"
    )
    parser.add_argument("-o", "--output", default="./temp/data", help="输出目录")
    parser.add_argument("-m", "--month", help="月账期 YYYYMM")
    parser.add_argument("-d", "--date", help="日账期 YYYYMMDD")
    parser.add_argument(
        "-n", "--name", help="数据集名称 (hr模式) / 配置名称 (config模式)"
    )
    parser.add_argument(
        "--qry-type", default="sql", help="查询类型 (direct/discover模式)"
    )
    parser.add_argument("-p", "--path", help="报告路径 (report/discover/direct模式)")
    parser.add_argument("-e", "--ename", help="图表名称 (report模式)")
    parser.add_argument(
        "--params", help="参数字符串: key1=value1,key2:type=value2 (direct/config模式)"
    )
    parser.add_argument("--file-type", default="xlsx", help="文件类型")
    parser.add_argument("--report-type", help="报告类型")
    parser.add_argument("--save", action="store_true", help="发现参数后保存到配置文件")
    parser.add_argument("--force", action="store_true", help="忽略本地已校验文件，强制重新下载")

    args = parser.parse_args()

    if args.output != "." and args.output != "downloads" and not os.path.exists(args.output):
        os.makedirs(args.output, exist_ok=True)

    if args.type in ["commission", "daily", "hr", "report", "sales_monthly"]:
        username, password = load_credentials()
        if not username or not password:
            Colors.error("[FAIL] 无法获取BI凭证，请先配置")
            Colors.info("[提示] 运行以下命令配置凭证:")
            Colors.info("  python -c \"from keyring_manager import *; save_credentials_keyring('your_user','your_pass')\"")
            sys.exit(1)

    if args.type == "commission":
        if args.month is None:
            args.month = get_last_month()
        run_commission_mode(args)
    elif args.type == "daily":
        if args.date is None:
            args.date = get_yesterday()
        run_daily_mode(args)
    elif args.type == "hr":
        if args.name is None:
            Colors.error("[FAIL] hr模式需要指定数据集名称 (-n/--name)")
            Colors.info("可用的数据集:")
            for key, info in HR_DATASETS.items():
                Colors.info(f"  - {key}: {info['name']}")
            sys.exit(1)
        run_hr_mode(args)
    elif args.type == "report":
        if args.path is None:
            Colors.error("[FAIL] report模式需要指定报告路径 (-p/--path)")
            sys.exit(1)
        if args.ename is None:
            Colors.error("[FAIL] report模式需要指定图表名称 (-e/--ename)")
            sys.exit(1)
        run_report_mode(args)
    elif args.type == "sales_monthly":
        if args.month is None:
            args.month = get_last_month()
        run_sales_monthly_mode(args)
    elif args.type == "discover":
        run_discover_mode(args)
    elif args.type == "config":
        if not args.name:
            Colors.error("[FAIL] config模式需要 -n 参数")
            sys.exit(1)
        run_config_mode(args)
    elif args.type == "direct":
        if not args.path:
            Colors.error("[FAIL] direct模式需要 --path 参数")
            sys.exit(1)
        run_direct_mode(args)

    try:
        from temp_file_manager import TempFileManager
        TempFileManager.cleanup_all_temp()
    except ImportError:
        pass


if __name__ == "__main__":
    main()
