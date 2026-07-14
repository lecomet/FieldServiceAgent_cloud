#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
import sys
import warnings
from decimal import Decimal, InvalidOperation
from pathlib import Path

try:
    import openpyxl
except ImportError as exc:
    raise SystemExit("缺少依赖 openpyxl，请使用 uv --with openpyxl 运行。") from exc

warnings.filterwarnings("ignore", message="Workbook contains no default style.*", category=UserWarning)

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")


BASE_COLUMNS = {"地市", "区县"}
PRODUCT_COLUMNS = ["移动", "终端宽带", "FTTR-H", "FTTR-B", "FTTR-H/B", "151", "天翼智屏"]


def project_root() -> Path:
    return Path(__file__).resolve().parents[4]


def latest_xlsx(directory: Path) -> Path | None:
    files = [path for path in directory.glob("*.xlsx") if path.is_file() and not path.name.startswith("~$")]
    if not files:
        return None
    return max(files, key=lambda path: path.stat().st_mtime)


def clean_header(value) -> str:
    return str(value or "").replace("\r", "").replace("\n", "").replace("\t", "").strip()


def format_fields(headers: list[str]) -> str:
    fields = [header for header in headers if header]
    return "、".join(fields) if fields else "无"


def validate_xlsx_file(path: Path) -> None:
    if path.suffix.lower() != ".xlsx":
        raise ValueError(f"{path} 文件格式不符合要求：只支持 .xlsx")
    with path.open("rb") as handle:
        signature = handle.read(4)
    if not signature.startswith(b"PK"):
        raise ValueError(f"{path} 不是标准 .xlsx 文件。请确认 BI 下载格式为 xlsx，不要把 CSV/HTML/XLS 改名为 .xlsx。")


def read_headers(path: Path, iterator) -> list[str]:
    try:
        raw_headers = next(iterator)
    except StopIteration as exc:
        raise ValueError(f"{path} 文件为空：第 1 行必须是字段名") from exc
    headers = [clean_header(value) for value in raw_headers]
    non_empty = [header for header in headers if header]
    duplicates = sorted({header for header in non_empty if non_empty.count(header) > 1})
    if duplicates:
        raise ValueError(f"{path} 第 1 行字段名重复: {', '.join(duplicates)}\n实际字段: {format_fields(headers)}")
    return headers


def require_columns(path: Path, indexes: dict[str, int], required: set[str], headers: list[str]) -> None:
    missing = sorted(required - set(indexes))
    if missing:
        raise ValueError(f"{path} 缺少字段: {', '.join(missing)}\n实际字段: {format_fields(headers)}")


def normalize_city(value: str) -> str:
    text = str(value or "").strip()
    return text[:-1] if text.endswith("市") else text


def text_match(value, query: str, normalize=None) -> bool:
    value_text = str(value or "").strip()
    query_text = str(query or "").strip()
    if normalize:
        value_text = normalize(value_text)
        query_text = normalize(query_text)
    return bool(value_text and query_text and (value_text == query_text or query_text in value_text or value_text in query_text))


def number(value) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    text = str(value).strip().replace(",", "")
    if not text:
        return Decimal("0")
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal("0")


def load_rows(path: Path, city: str | None, county: str | None) -> tuple[list[str], list[dict]]:
    validate_xlsx_file(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    iterator = ws.iter_rows(values_only=True)
    try:
        headers = read_headers(path, iterator)
        indexes = {header: idx for idx, header in enumerate(headers) if header}
        require_columns(path, indexes, BASE_COLUMNS, headers)

        available_products = [column for column in PRODUCT_COLUMNS if column in indexes]
        if not available_products:
            raise ValueError(f"{path} 缺少可计算破零率的发展量字段: {', '.join(PRODUCT_COLUMNS)}\n实际字段: {format_fields(headers)}")

        rows = []
        for raw in iterator:
            if city and not text_match(raw[indexes["地市"]], city, normalize_city):
                continue
            if county and not text_match(raw[indexes["区县"]], county):
                continue
            rows.append({header: raw[idx] if idx < len(raw) else None for header, idx in indexes.items()})
    finally:
        wb.close()
    return available_products, rows


def normalize_product_text(value: str) -> str:
    return str(value or "").upper().replace(" ", "").replace("－", "-").replace("_", "")


def detect_product(question: str, explicit_product: str | None) -> str | None:
    raw = explicit_product or question or ""
    compact = normalize_product_text(raw)
    if explicit_product in PRODUCT_COLUMNS:
        return explicit_product
    if "FTTR-H/B" in compact or "FTTRH/B" in compact or "FTTR-HB" in compact or "FTTRHB" in compact:
        return "FTTR-H/B"
    if "FTTR-H" in compact or "FTTRH" in compact:
        return "FTTR-H"
    if "FTTR-B" in compact or "FTTRB" in compact:
        return "FTTR-B"
    if "FTTR" in compact:
        return "FTTR-H/B"
    if "151" in compact:
        return "151"
    if "天翼智屏" in raw or "智屏" in raw:
        return "天翼智屏"
    if "终端宽带" in raw or "宽带" in raw:
        return "终端宽带"
    if "移动" in raw:
        return "移动"
    return None


def scope_text(city: str | None, county: str | None) -> str:
    if county:
        return f"区县={county}"
    if city:
        return f"地市={city}"
    return "全省"


def main() -> None:
    parser = argparse.ArgumentParser(description="装维正式人员破零率计算")
    parser.add_argument("--question", default="", help="用户问题原文")
    parser.add_argument("--product", help="产品字段：移动/终端宽带/FTTR-H/FTTR-B/FTTR-H/B/151/天翼智屏")
    parser.add_argument("--city", help="地市名称")
    parser.add_argument("--county", help="区县名称")
    parser.add_argument("--official-dir", default="temp/data/official")
    args = parser.parse_args()

    root = project_root()
    source_file = latest_xlsx(root / args.official_dir)
    if not source_file:
        raise SystemExit(f"未找到正式人员 Excel: {root / args.official_dir}")

    available_products, rows = load_rows(source_file, args.city, args.county)
    if not rows:
        raise SystemExit("未找到匹配范围内的正式人员数据。")

    product = detect_product(args.question, args.product)
    if product and product not in available_products:
        raise SystemExit(f"来源文件缺少产品字段: {product}")

    if product:
        zero_break_rows = [row for row in rows if number(row.get(product)) > 0]
        product_label = product
    else:
        zero_break_rows = [row for row in rows if any(number(row.get(column)) > 0 for column in available_products)]
        product_label = "任一发展量"

    total = len(rows)
    zero_break_count = len(zero_break_rows)
    ratio = Decimal(zero_break_count) / Decimal(total) * Decimal("100") if total else Decimal("0")

    print("=== 装维正式人员破零率 ===")
    print(f"范围: {scope_text(args.city, args.county)}")
    print(f"来源文件: {source_file}")
    print(f"产品字段: {product_label}")
    print("口径: 破零人数 = 产品发展量 > 0 的正式人员数；破零率 = 破零人数 / 匹配范围内正式人员数")
    print(f"样本人数: {total}")
    print(f"破零人数: {zero_break_count}")
    print(f"未破零人数: {total - zero_break_count}")
    print(f"破零率: {round(float(ratio), 2)}%")


if __name__ == "__main__":
    main()
