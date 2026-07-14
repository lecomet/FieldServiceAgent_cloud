#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
import sys
import warnings
from decimal import Decimal, InvalidOperation
from pathlib import Path

try:
    import openpyxl
except ImportError as exc:
    raise SystemExit("缺少依赖 openpyxl，请使用 uv --with openpyxl 运行。") from exc

warnings.filterwarnings("ignore", message="Workbook contains no default style.*", category=UserWarning)

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")


REQUIRED_COLUMNS = {"地市", "区县", "用工属性", "新老装维", "原始积分"}


def project_root() -> Path:
    return Path(__file__).resolve().parents[4]


def latest_xlsx(directory: Path) -> Path | None:
    files = [path for path in directory.glob("*.xlsx") if path.is_file() and not path.name.startswith("~$")]
    if not files:
        return None
    return max(files, key=lambda path: path.stat().st_mtime)


def clean_header(value) -> str:
    return str(value or "").replace("\r", "").replace("\n", "").replace("\t", "").strip()


def format_fields(headers: list[str]) -> str:
    fields = [header for header in headers if header]
    return "、".join(fields) if fields else "无"


def validate_xlsx_file(path: Path) -> None:
    if path.suffix.lower() != ".xlsx":
        raise ValueError(f"{path} 文件格式不符合要求：只支持 .xlsx")
    with path.open("rb") as handle:
        signature = handle.read(4)
    if not signature.startswith(b"PK"):
        raise ValueError(f"{path} 不是标准 .xlsx 文件。请确认 BI 下载格式为 xlsx，不要把 CSV/HTML/XLS 改名为 .xlsx。")


def read_headers(path: Path, iterator) -> list[str]:
    try:
        raw_headers = next(iterator)
    except StopIteration as exc:
        raise ValueError(f"{path} 文件为空：第 1 行必须是字段名") from exc
    headers = [clean_header(value) for value in raw_headers]
    non_empty = [header for header in headers if header]
    duplicates = sorted({header for header in non_empty if non_empty.count(header) > 1})
    if duplicates:
        raise ValueError(f"{path} 第 1 行字段名重复: {', '.join(duplicates)}\n实际字段: {format_fields(headers)}")
    return headers


def require_columns(path: Path, indexes: dict[str, int], required: set[str], headers: list[str]) -> None:
    missing = sorted(required - set(indexes))
    if missing:
        raise ValueError(f"{path} 缺少字段: {', '.join(missing)}\n实际字段: {format_fields(headers)}")


def normalize_city(value: str) -> str:
    text = str(value or "").strip()
    return text[:-1] if text.endswith("市") else text


def text_match(value, query: str, normalize=None) -> bool:
    value_text = str(value or "").strip()
    query_text = str(query or "").strip()
    if normalize:
        value_text = normalize(value_text)
        query_text = normalize(query_text)
    return bool(value_text and query_text and (value_text == query_text or query_text in value_text or value_text in query_text))


def number(value) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    text = str(value).strip().replace(",", "")
    if not text:
        return Decimal("0")
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal("0")


def load_rows(path: Path, city: str | None, county: str | None) -> list[dict]:
    validate_xlsx_file(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    iterator = ws.iter_rows(values_only=True)
    try:
        headers = read_headers(path, iterator)
        indexes = {header: idx for idx, header in enumerate(headers) if header}
        require_columns(path, indexes, REQUIRED_COLUMNS, headers)

        rows = []
        for raw in iterator:
            if city and not text_match(raw[indexes["地市"]], city, normalize_city):
                continue
            if county and not text_match(raw[indexes["区县"]], county):
                continue
            rows.append({header: raw[idx] if idx < len(raw) else None for header, idx in indexes.items()})
    finally:
        wb.close()
    return rows


def is_cp_staff(row: dict) -> bool:
    attr = str(row.get("用工属性") or "").strip()
    old_new = str(row.get("新老装维") or "").strip()
    return old_new in {"新装维", "老装维"} and attr != "装维门店" and ("装维" in attr or "智家工程师" in attr)


def scope_text(city: str | None, county: str | None) -> str:
    if county:
        return f"区县={county}"
    if city:
        return f"地市={city}"
    return "全省"


def main() -> None:
    parser = argparse.ArgumentParser(description="装维 CP 积分指标计算")
    parser.add_argument("--question", default="", help="用户问题原文")
    parser.add_argument("--city", help="地市名称")
    parser.add_argument("--county", help="区县名称")
    parser.add_argument("--official-dir", default="temp/data/official")
    args = parser.parse_args()

    root = project_root()
    source_file = latest_xlsx(root / args.official_dir)
    if not source_file:
        raise SystemExit(f"未找到正式人员 Excel: {root / args.official_dir}")

    rows = load_rows(source_file, args.city, args.county)
    if not rows:
        raise SystemExit("未找到匹配范围内的正式人员数据。")

    cp_rows = [row for row in rows if is_cp_staff(row)]
    cp_score = sum((number(row.get("原始积分")) for row in cp_rows), Decimal("0"))
    total_score = sum((number(row.get("原始积分")) for row in rows), Decimal("0"))
    ratio = (cp_score / total_score * Decimal("100")) if total_score else Decimal("0")

    print("=== 装维 CP 积分指标 ===")
    print(f"范围: {scope_text(args.city, args.county)}")
    print(f"来源文件: {source_file}")
    print("CP人员口径: 正式人员 + 用工属性含“装维”或“智家工程师” + 用工属性不是“装维门店” + 新老装维为新装维/老装维")
    print("CP积分口径: 非开店正式人员原始积分")
    print("CP积分占比口径: CP积分 / 全部正式人员原始积分，不包含实习人员")
    print(f"样本人数: {len(rows)}")
    print(f"CP人员数: {len(cp_rows)}")
    print(f"CP原始积分: {round(float(cp_score), 2)}")
    print(f"全部正式人员原始积分: {round(float(total_score), 2)}")
    print(f"CP积分占比: {round(float(ratio), 2)}%")


if __name__ == "__main__":
    main()
