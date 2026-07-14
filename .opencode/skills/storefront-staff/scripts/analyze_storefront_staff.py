#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
import sys
import warnings
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

try:
    import openpyxl
except ImportError as exc:
    raise SystemExit("缺少依赖 openpyxl，请使用 uv --with openpyxl 运行。") from exc

warnings.filterwarnings("ignore", message="Workbook contains no default style.*", category=UserWarning)

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")


REQUIRED_COLUMNS = {"地市", "区县", "装维姓名", "工号", "用工属性", "新老装维", "原始积分"}
DISPLAY_COLUMNS = ["地市", "区县", "装维姓名", "工号", "用工属性", "新老装维", "原始积分"]
DEFAULT_MAX_OUTPUT_ROWS = 20


def project_root() -> Path:
    return Path(__file__).resolve().parents[4]


def latest_xlsx(directory: Path) -> Path | None:
    files = [path for path in directory.glob("*.xlsx") if path.is_file() and not path.name.startswith("~$")]
    if not files:
        return None
    return max(files, key=lambda path: path.stat().st_mtime)


def clean_header(value) -> str:
    return str(value or "").replace("\r", "").replace("\n", "").replace("\t", "").strip()


def format_fields(headers: list[str]) -> str:
    fields = [header for header in headers if header]
    return "、".join(fields) if fields else "无"


def validate_xlsx_file(path: Path) -> None:
    if path.suffix.lower() != ".xlsx":
        raise ValueError(f"{path} 文件格式不符合要求：只支持 .xlsx")
    with path.open("rb") as handle:
        signature = handle.read(4)
    if not signature.startswith(b"PK"):
        raise ValueError(f"{path} 不是标准 .xlsx 文件。请确认 BI 下载格式为 xlsx，不要把 CSV/HTML/XLS 改名为 .xlsx。")


def read_headers(path: Path, iterator) -> list[str]:
    try:
        raw_headers = next(iterator)
    except StopIteration as exc:
        raise ValueError(f"{path} 文件为空：第 1 行必须是字段名") from exc
    headers = [clean_header(value) for value in raw_headers]
    non_empty = [header for header in headers if header]
    duplicates = sorted({header for header in non_empty if non_empty.count(header) > 1})
    if duplicates:
        raise ValueError(f"{path} 第 1 行字段名重复: {', '.join(duplicates)}\n实际字段: {format_fields(headers)}")
    return headers


def require_columns(path: Path, indexes: dict[str, int], required: set[str], headers: list[str]) -> None:
    missing = sorted(required - set(indexes))
    if missing:
        raise ValueError(f"{path} 缺少字段: {', '.join(missing)}\n实际字段: {format_fields(headers)}")


def normalize_city(value: str) -> str:
    text = str(value or "").strip()
    return text[:-1] if text.endswith("市") else text


def text_match(value, query: str, normalize=None) -> bool:
    value_text = str(value or "").strip()
    query_text = str(query or "").strip()
    if normalize:
        value_text = normalize(value_text)
        query_text = normalize(query_text)
    return bool(value_text and query_text and (value_text == query_text or query_text in value_text or value_text in query_text))


def number(value) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    text = str(value).strip().replace(",", "")
    if not text:
        return Decimal("0")
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal("0")


def load_rows(path: Path, city: str | None, county: str | None) -> tuple[list[str], list[dict]]:
    validate_xlsx_file(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    iterator = ws.iter_rows(values_only=True)
    try:
        headers = read_headers(path, iterator)
        indexes = {header: idx for idx, header in enumerate(headers) if header}
        require_columns(path, indexes, REQUIRED_COLUMNS, headers)

        rows = []
        for raw in iterator:
            if city and not text_match(raw[indexes["地市"]], city, normalize_city):
                continue
            if county and not text_match(raw[indexes["区县"]], county):
                continue
            rows.append({header: raw[idx] if idx < len(raw) else None for header, idx in indexes.items()})
    finally:
        wb.close()
    return headers, rows


def is_open_shop(row: dict) -> bool:
    return str(row.get("用工属性") or "").strip() == "装维门店" and str(row.get("新老装维") or "").strip() in {"新装维", "老装维"}


def is_non_open_shop(row: dict) -> bool:
    attr = str(row.get("用工属性") or "").strip()
    old_new = str(row.get("新老装维") or "").strip()
    return old_new in {"新装维", "老装维"} and attr != "装维门店" and ("装维" in attr or "智家工程师" in attr)


def detect_metric(question: str, explicit_metric: str | None) -> str:
    if explicit_metric:
        return explicit_metric
    text = question or ""
    wants_list = any(word in text for word in ["清单", "名单", "明细"])
    if "非开店" in text and wants_list:
        return "non_open_shop_list"
    if "开店" in text and wants_list:
        return "open_shop_list"
    if "开店" in text and "人均积分" in text:
        return "open_shop_avg_score"
    if "非开店" in text:
        return "non_open_shop_count"
    if "开店" in text:
        return "open_shop_count"
    return "unknown"


def scope_text(city: str | None, county: str | None) -> str:
    if county:
        return f"区县={county}"
    if city:
        return f"地市={city}"
    return "全省"


def scope_token(city: str | None, county: str | None) -> str:
    if county:
        return f"county_{county}"
    if city:
        return f"city_{normalize_city(city)}"
    return "province"


def markdown_table(rows: list[dict]) -> str:
    lines = [
        "| " + " | ".join(DISPLAY_COLUMNS) + " |",
        "| " + " | ".join("---" for _ in DISPLAY_COLUMNS) + " |",
    ]
    for row in rows:
        lines.append("| " + " | ".join(str(row.get(col) or "") for col in DISPLAY_COLUMNS) + " |")
    return "\n".join(lines)


def write_list(output_path: Path, rows: list[dict], source_file: Path, metric: str) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "人员清单"
    ws.append(DISPLAY_COLUMNS)
    for row in rows:
        ws.append([row.get(col) for col in DISPLAY_COLUMNS])

    meta = wb.create_sheet("来源")
    meta.append(["项目", "值"])
    meta.append(["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    meta.append(["来源文件", str(source_file)])
    meta.append(["指标", metric])
    meta.append(["行数", len(rows)])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(output_path)
        return output_path
    except PermissionError:
        fallback = output_path.with_name(f"{output_path.stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{output_path.suffix}")
        wb.save(fallback)
        return fallback


def print_metric_header(source_file: Path, city: str | None, county: str | None) -> None:
    print("=== 装维开店/非开店人员指标 ===")
    print(f"范围: {scope_text(city, county)}")
    print(f"来源文件: {source_file}")
    print("开店员工口径: 正式员工 + 用工属性为“装维门店” + 新老装维为新装维/老装维")
    print("非开店员工口径: 正式员工 + 用工属性含“装维”或“智家工程师” + 用工属性不是“装维门店” + 新老装维为新装维/老装维")


def main() -> None:
    parser = argparse.ArgumentParser(description="装维开店/非开店人员指标计算")
    parser.add_argument("--question", default="", help="用户问题原文")
    parser.add_argument("--metric", help="open_shop_count/non_open_shop_count/open_shop_avg_score/open_shop_list/non_open_shop_list")
    parser.add_argument("--city", help="地市名称")
    parser.add_argument("--county", help="区县名称")
    parser.add_argument("--official-dir", default="temp/data/official")
    parser.add_argument("--max-output", type=int, default=DEFAULT_MAX_OUTPUT_ROWS)
    args = parser.parse_args()

    root = project_root()
    source_file = latest_xlsx(root / args.official_dir)
    if not source_file:
        raise SystemExit(f"未找到正式人员 Excel: {root / args.official_dir}")

    _, rows = load_rows(source_file, args.city, args.county)
    if not rows:
        raise SystemExit("未找到匹配范围内的正式人员数据。")

    metric = detect_metric(args.question, args.metric)
    open_rows = [row for row in rows if is_open_shop(row)]
    non_open_rows = [row for row in rows if is_non_open_shop(row)]

    print_metric_header(source_file, args.city, args.county)
    if metric == "open_shop_count":
        print("指标: 开店人数")
        print(f"结果: {len(open_rows)}")
        print(f"样本人数: {len(rows)}")
        return

    if metric == "non_open_shop_count":
        print("指标: 非开店人数")
        print(f"结果: {len(non_open_rows)}")
        print(f"样本人数: {len(rows)}")
        return

    if metric == "open_shop_avg_score":
        score_sum = sum((number(row.get("原始积分")) for row in open_rows), Decimal("0"))
        avg_score = score_sum / Decimal(len(open_rows)) if open_rows else Decimal("0")
        print("指标: 开店人员人均积分")
        print(f"开店人数: {len(open_rows)}")
        print(f"原始积分合计: {round(float(score_sum), 2)}")
        print(f"结果: {round(float(avg_score), 2)}")
        return

    if metric in {"open_shop_list", "non_open_shop_list"}:
        selected = open_rows if metric == "open_shop_list" else non_open_rows
        label = "开店人员" if metric == "open_shop_list" else "非开店人员"
        output = root / "output" / f"storefront_{scope_token(args.city, args.county)}_{label}.xlsx"
        saved = write_list(output, selected, source_file, label)
        print(f"指标: {label}清单")
        print(f"清单文件: {saved}")
        print(f"明细行数: {len(selected)}")
        if len(selected) > args.max_output:
            print(f"明细行数超过 {args.max_output} 行，不在消息栏罗列；请打开 Excel 查看完整清单。")
            return
        if selected:
            print(markdown_table(selected))
        return

    print("状态: 未识别问题")
    print("提示: 请明确为开店人数、非开店人数、开店人员清单、非开店人员清单或开店人员人均积分。")


if __name__ == "__main__":
    main()
