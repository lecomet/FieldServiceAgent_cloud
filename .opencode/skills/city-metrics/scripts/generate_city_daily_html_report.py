#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
import html
import re
import sys
import warnings
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path

try:
    import openpyxl
except ImportError as exc:
    raise SystemExit("缺少依赖 openpyxl，请使用 uv --with openpyxl 运行。") from exc

warnings.filterwarnings("ignore", message="Workbook contains no default style.*", category=UserWarning)

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")


REQUIRED_COLUMNS = {"地市", "区县"}
RATIO_PAIRS = {"装维占比": ("装维发展量", "全量发展量")}
FOCUS_LIMIT = 5
FOCUS_CANDIDATES = [
    ("151", ["151"], "number"),
    ("FTTR-H/B", ["FTTR-H/B", "FTTR"], "number"),
    ("人均价值积分", ["人均价值积分", "人均原始积分", "原始积分"], "number"),
]
DETAIL_PREFERRED = [
    "区县",
    "移动",
    "宽带",
    "终端宽带",
    "FTTR-H",
    "FTTR-B",
    "FTTR-H/B",
    "151",
    "装维发展量",
    "全量发展量",
    "装维占比",
    "发展积分",
    "运营积分",
    "价值积分",
    "原始积分",
    "人均价值积分",
    "人均原始积分",
]


def project_root() -> Path:
    return Path(__file__).resolve().parents[4]


def latest_xlsx(directory: Path) -> Path | None:
    files = [path for path in directory.glob("*.xlsx") if path.is_file() and not path.name.startswith("~$")]
    if not files:
        return None
    return max(files, key=lambda path: path.stat().st_mtime)


def clean_header(value) -> str:
    return str(value or "").replace("\r", "").replace("\n", "").replace("\t", "").strip()


def format_fields(headers: list[str]) -> str:
    fields = [header for header in headers if header]
    return "、".join(fields) if fields else "无"


def validate_xlsx_file(path: Path) -> None:
    if path.suffix.lower() != ".xlsx":
        raise ValueError(f"{path} 文件格式不符合要求：只支持 .xlsx")
    with path.open("rb") as handle:
        signature = handle.read(4)
    if not signature.startswith(b"PK"):
        raise ValueError(f"{path} 不是标准 .xlsx 文件。请确认 BI 下载格式为 xlsx，不要把 CSV/HTML/XLS 改名为 .xlsx。")


def read_headers(path: Path, iterator) -> list[str]:
    try:
        raw_headers = next(iterator)
    except StopIteration as exc:
        raise ValueError(f"{path} 文件为空：第 1 行必须是字段名") from exc
    headers = [clean_header(value) for value in raw_headers]
    non_empty = [header for header in headers if header]
    duplicates = sorted({header for header in non_empty if non_empty.count(header) > 1})
    if duplicates:
        raise ValueError(f"{path} 第 1 行字段名重复: {', '.join(duplicates)}\n实际字段: {format_fields(headers)}")
    return headers


def require_columns(path: Path, indexes: dict[str, int], required: set[str], headers: list[str]) -> None:
    missing = sorted(required - set(indexes))
    if missing:
        raise ValueError(f"{path} 缺少字段: {', '.join(missing)}\n实际字段: {format_fields(headers)}")


def normalize_city(value) -> str:
    text = str(value or "").strip()
    return text[:-1] if text.endswith("市") else text


def normalize_county(value) -> str:
    return str(value or "").strip()


def text_match(value, query: str, normalize=None) -> bool:
    value_text = str(value or "").strip()
    query_text = str(query or "").strip()
    if normalize:
        value_text = normalize(value_text)
        query_text = normalize(query_text)
    return bool(value_text and query_text and (value_text == query_text or query_text in value_text or value_text in query_text))


def decimal_value(value) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = str(value).strip().replace(",", "").replace("%", "")
    if not text:
        return Decimal("0")
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal("0")


def parse_dates(path: Path) -> tuple[str, str]:
    match = re.search(r"_(\d{8})_(\d{6})_", path.name)
    if not match:
        return "", ""
    return match.group(1), match.group(2)


def format_day(day: str) -> str:
    if len(day) != 8:
        return day
    return f"{day[:4]}-{day[4:6]}-{day[6:8]}"


def escape(value) -> str:
    return html.escape(str(value or ""), quote=True)


def format_number(value, kind: str = "number") -> str:
    number = decimal_value(value)
    if kind == "ratio":
        pct = (number * Decimal("100")).quantize(Decimal("0.1"), rounding=ROUND_HALF_UP)
        return f"{pct}%"
    if number == number.to_integral_value():
        return f"{int(number):,}"
    rounded = number.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    return f"{rounded:,}"


def load_rows(path: Path, city: str) -> tuple[list[str], list[dict]]:
    validate_xlsx_file(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    iterator = ws.iter_rows(values_only=True)
    try:
        headers = read_headers(path, iterator)
        indexes = {header: idx for idx, header in enumerate(headers) if header}
        require_columns(path, indexes, REQUIRED_COLUMNS, headers)

        rows = []
        for raw in iterator:
            if not raw or not any(value is not None and value != "" for value in raw):
                continue
            if not text_match(raw[indexes["地市"]], city, normalize_city):
                continue
            row = {header: raw[idx] if idx < len(raw) else None for header, idx in indexes.items()}
            row["地市"] = normalize_city(row.get("地市"))
            row["区县"] = normalize_county(row.get("区县"))
            if row.get("区县"):
                rows.append(row)
    finally:
        wb.close()
    return headers, rows


def aggregate_rows(headers: list[str], rows: list[dict]) -> list[dict]:
    grouped: dict[tuple[str, str], dict] = {}
    for row in rows:
        city = normalize_city(row.get("地市"))
        county = normalize_county(row.get("区县"))
        if not city or not county:
            continue
        target = grouped.setdefault((city, county), {header: None for header in headers})
        target["地市"] = city
        target["区县"] = county
        for header in headers:
            if header in {"地市", "区县"} or header in RATIO_PAIRS:
                continue
            target[header] = decimal_value(target.get(header)) + decimal_value(row.get(header))

    for row in grouped.values():
        for ratio_col, (num_col, den_col) in RATIO_PAIRS.items():
            if ratio_col not in headers or num_col not in headers or den_col not in headers:
                continue
            denominator = decimal_value(row.get(den_col))
            row[ratio_col] = decimal_value(row.get(num_col)) / denominator if denominator else Decimal("0")
    return list(grouped.values())


def resolve_focus_metrics(headers: list[str]) -> tuple[list[dict], list[str]]:
    metrics = []
    missing = []
    for title, candidates, kind in FOCUS_CANDIDATES:
        key = next((candidate for candidate in candidates if candidate in headers), "")
        if key:
            metrics.append({"title": title, "key": key, "kind": kind})
        else:
            missing.append(title)
    return metrics, missing


def total_value(rows: list[dict], key: str, kind: str) -> str:
    if key.startswith("人均"):
        values = [decimal_value(row.get(key)) for row in rows]
        if not values:
            return "0"
        return format_number(sum(values, Decimal("0")) / Decimal(len(values)), kind)
    return format_number(sum((decimal_value(row.get(key)) for row in rows), Decimal("0")), kind)


def rank_items(rows: list[dict], key: str, reverse: bool) -> list[tuple[int, str, Decimal]]:
    values = [(normalize_county(row.get("区县")), decimal_value(row.get(key))) for row in rows if normalize_county(row.get("区县"))]
    values = sorted(values, key=lambda item: ((-item[1] if reverse else item[1]), item[0]))
    return [(rank, county, value) for rank, (county, value) in enumerate(values[:FOCUS_LIMIT], start=1)]


def rank_classes(rows: list[dict], metrics: list[dict]) -> dict[tuple[int, str], str]:
    classes: dict[tuple[int, str], str] = {}
    for metric in metrics:
        key = metric["key"]
        values = [(idx, decimal_value(row.get(key)), normalize_county(row.get("区县"))) for idx, row in enumerate(rows)]
        if not values:
            continue
        unique_values = {value for _, value, _ in values}
        if len(unique_values) <= 1:
            continue
        top = sorted(values, key=lambda item: (-item[1], item[2]))[:FOCUS_LIMIT]
        bottom = sorted(values, key=lambda item: (item[1], item[2]))[:FOCUS_LIMIT]
        for idx, _, _ in bottom:
            classes[(idx, key)] = "rank-low"
        for idx, _, _ in top:
            classes[(idx, key)] = "rank-high"
    return classes


def render_rank_rows(items: list[tuple[int, str, Decimal]], kind: str) -> str:
    if not items:
        return '<div class="rank-empty">暂无数据</div>'
    rows = []
    for rank, county, value in items:
        rows.append(
            f"""
            <div class="rank-row">
              <div class="rank-no">{rank}</div>
              <div class="rank-city">{escape(county)}</div>
              <div class="rank-value">{escape(format_number(value, kind))}</div>
            </div>
            """
        )
    return "".join(rows)


def render_focus_card(rows: list[dict], metric: dict) -> str:
    key = metric["key"]
    title = metric["title"]
    kind = metric["kind"]
    return f"""
      <section class="focus-card">
        <div class="focus-head">
          <div>
            <h2>{escape(title)}</h2>
            <div class="focus-subtitle">区县 TOP 5 / LAST 5</div>
          </div>
          <div class="focus-total">
            <span>地市汇总</span>
            <strong>{total_value(rows, key, kind)}</strong>
          </div>
        </div>
        <div class="rank-columns">
          <div class="rank-panel rank-panel-top">
            <div class="rank-title">TOP 5</div>
            {render_rank_rows(rank_items(rows, key, reverse=True), kind)}
          </div>
          <div class="rank-panel rank-panel-last">
            <div class="rank-title">LAST 5</div>
            {render_rank_rows(rank_items(rows, key, reverse=False), kind)}
          </div>
        </div>
      </section>
    """


def render_focus_grid(rows: list[dict], metrics: list[dict], missing: list[str]) -> str:
    cards = "".join(render_focus_card(rows, metric) for metric in metrics)
    warning = ""
    if missing:
        warning = f'<div class="missing-warning">缺少重点字段：{escape("、".join(missing))}</div>'
    return f'<div class="focus-grid">{cards}</div>{warning}'


def chart_items(rows: list[dict], key: str, reverse: bool) -> list[tuple[str, Decimal]]:
    values = [
        (normalize_county(row.get("区县")), decimal_value(row.get(key)))
        for row in rows
        if normalize_county(row.get("区县"))
    ]
    return sorted(values, key=lambda item: ((-item[1] if reverse else item[1]), item[0]))[:FOCUS_LIMIT]


def render_bar_rows(items: list[tuple[str, Decimal]], max_value: Decimal, kind: str, fill_class: str) -> str:
    if not items:
        return ""
    bars = []
    for county, value in items:
        width = max(8, int((value / max_value) * Decimal("100")))
        bars.append(
            f"""
            <div class="bar-row">
              <div class="bar-label">{escape(county)}</div>
              <div class="bar-track"><div class="bar-fill {fill_class}" style="width:{width}%"></div></div>
              <div class="bar-value">{escape(format_number(value, kind))}</div>
            </div>
            """
        )
    return "".join(bars)


def render_chart(title: str, rows: list[dict], key: str, kind: str = "number") -> str:
    top = chart_items(rows, key, reverse=True)
    last = chart_items(rows, key, reverse=False)
    combined = top + last
    if not combined:
        return ""
    max_value = max((value for _, value in combined), default=Decimal("0")) or Decimal("1")
    return f"""
      <section class="chart-card">
        <h2>{escape(title)}</h2>
        <div class="chart-columns">
          <div class="chart-panel">
            <div class="chart-label chart-label-top">TOP 5</div>
            {render_bar_rows(top, max_value, kind, "bar-fill-top")}
          </div>
          <div class="chart-panel">
            <div class="chart-label chart-label-last">LAST 5</div>
            {render_bar_rows(last, max_value, kind, "bar-fill-last")}
          </div>
        </div>
      </section>
    """


def render_focus_charts(rows: list[dict], metrics: list[dict], missing: list[str]) -> str:
    charts = "".join(render_chart(metric["title"], rows, metric["key"], metric["kind"]) for metric in metrics)
    warning = ""
    if missing:
        warning = f'<div class="missing-warning">缺少重点字段：{escape("、".join(missing))}</div>'
    return f"""
      <section class="chart-section">
        <div class="section-title">
          <h2>重点指标</h2>
          <span>151装维量 / FTTR-H/B装维量 / 人均价值积分，均展示 TOP 5 和 LAST 5</span>
        </div>
        <div class="charts">{charts}</div>
        {warning}
      </section>
    """


def detail_columns(headers: list[str], metrics: list[dict]) -> list[dict]:
    selected = []
    ordered = ["区县"] + [header for header in headers if header not in {"地市", "区县"}]
    for key in ordered:
        if key in headers and key not in selected:
            selected.append(key)
    for metric in metrics:
        if metric["key"] not in selected:
            selected.append(metric["key"])
    return [
        {"key": key, "title": key, "kind": "ratio" if "占比" in key else "number" if key != "区县" else "text"}
        for key in selected
    ]


def render_detail_table(rows: list[dict], columns: list[dict], metrics: list[dict]) -> str:
    classes = rank_classes(rows, metrics)
    headers = "".join(f'<th>{escape(col["title"])}</th>' for col in columns)
    body_rows = []
    for idx, row in enumerate(rows):
        cells = []
        for col in columns:
            key = col["key"]
            kind = col["kind"]
            if kind == "text":
                cells.append(f'<td class="fixed">{escape(row.get(key))}</td>')
                continue
            cell_class = "num"
            rank_class = classes.get((idx, key))
            if rank_class:
                cell_class += f" {rank_class}"
            cells.append(f'<td class="{cell_class}">{escape(format_number(row.get(key), kind))}</td>')
        body_rows.append(f"<tr>{''.join(cells)}</tr>")
    return f"""
      <table class="report-table">
        <thead>
          <tr>{headers}</tr>
        </thead>
        <tbody>
          {''.join(body_rows)}
        </tbody>
      </table>
    """


def render_html(source: Path, city: str, headers: list[str], rows: list[dict]) -> str:
    acct_day, month_id = parse_dates(source)
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    title_date = format_day(acct_day) if acct_day else "最新"
    metrics, missing = resolve_focus_metrics(headers)
    focus_charts = render_focus_charts(rows, metrics, missing)
    table = render_detail_table(rows, detail_columns(headers, metrics), metrics)
    city_name = normalize_city(city)

    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{escape(city_name)}地市日通报_{escape(acct_day or "latest")}</title>
  <style>
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      background: #eef2f6;
      color: #172033;
      font-family: "Microsoft YaHei", "PingFang SC", Arial, sans-serif;
    }}
    .page {{
      width: 1840px;
      margin: 0 auto;
      padding: 30px 34px 34px;
      background: #f7f9fc;
    }}
    .title-row {{
      display: flex;
      align-items: flex-end;
      justify-content: space-between;
      gap: 24px;
      margin-bottom: 18px;
      border-bottom: 3px solid #1d4e89;
      padding-bottom: 16px;
    }}
    h1 {{
      margin: 0;
      color: #102a43;
      font-size: 34px;
      line-height: 1.15;
      letter-spacing: 0;
      font-weight: 800;
    }}
    .subtitle {{
      margin-top: 8px;
      font-size: 16px;
      color: #52606d;
    }}
    .meta {{
      text-align: right;
      font-size: 14px;
      color: #334e68;
      line-height: 1.8;
      white-space: nowrap;
    }}
    .focus-grid {{
      display: grid;
      grid-template-columns: repeat(3, 1fr);
      gap: 14px;
      margin-bottom: 18px;
    }}
    .focus-card {{
      background: #ffffff;
      border: 1px solid #cbd5e1;
      border-top: 5px solid #1d4e89;
      border-radius: 6px;
      padding: 16px;
      min-height: 430px;
    }}
    .focus-head {{
      display: flex;
      justify-content: space-between;
      align-items: flex-start;
      gap: 14px;
      margin-bottom: 14px;
      padding-bottom: 12px;
      border-bottom: 1px solid #d9e2ec;
    }}
    .focus-card h2 {{
      margin: 0;
      color: #102a43;
      font-size: 26px;
      line-height: 1.2;
      letter-spacing: 0;
    }}
    .focus-subtitle {{
      margin-top: 5px;
      color: #52606d;
      font-size: 14px;
    }}
    .focus-total {{
      min-width: 116px;
      text-align: right;
    }}
    .focus-total span {{
      display: block;
      color: #52606d;
      font-size: 13px;
      margin-bottom: 5px;
    }}
    .focus-total strong {{
      display: block;
      color: #102a43;
      font-size: 26px;
      line-height: 1.1;
      font-variant-numeric: tabular-nums;
    }}
    .rank-columns {{
      display: grid;
      grid-template-columns: 1fr 1fr;
      gap: 12px;
    }}
    .rank-panel {{
      border: 1px solid #d9e2ec;
      border-radius: 6px;
      overflow: hidden;
      background: #f8fafc;
    }}
    .rank-title {{
      padding: 9px 10px;
      color: #ffffff;
      font-size: 15px;
      font-weight: 800;
      letter-spacing: 0;
    }}
    .rank-panel-top .rank-title {{
      background: #137333;
    }}
    .rank-panel-last .rank-title {{
      background: #b42318;
    }}
    .rank-row {{
      display: grid;
      grid-template-columns: 34px 1fr 88px;
      align-items: center;
      min-height: 44px;
      padding: 8px 9px;
      border-top: 1px solid #e2e8f0;
      background: #ffffff;
      gap: 8px;
      font-size: 15px;
    }}
    .rank-no {{
      width: 26px;
      height: 26px;
      line-height: 26px;
      border-radius: 50%;
      background: #e2e8f0;
      color: #172033;
      text-align: center;
      font-weight: 800;
      font-variant-numeric: tabular-nums;
    }}
    .rank-city {{
      color: #172033;
      font-weight: 800;
      white-space: nowrap;
      overflow: hidden;
      text-overflow: ellipsis;
    }}
    .rank-value {{
      color: #102a43;
      font-weight: 900;
      text-align: right;
      font-variant-numeric: tabular-nums;
    }}
    .missing-warning {{
      margin: -6px 0 14px;
      padding: 10px 12px;
      border: 1px solid #f8b4a6;
      border-radius: 6px;
      background: #fff4f2;
      color: #b42318;
      font-size: 14px;
      font-weight: 700;
    }}
    .table-wrap {{
      background: #ffffff;
      border: 1px solid #cbd5e1;
      border-radius: 6px;
      overflow: hidden;
      margin-bottom: 16px;
    }}
    .report-table {{
      width: 100%;
      border-collapse: collapse;
      table-layout: fixed;
      font-size: 15px;
      line-height: 1.25;
    }}
    .report-table th {{
      border: 1px solid #94a3b8;
      background: #dbeafe;
      color: #0f2747;
      padding: 9px 6px;
      text-align: center;
      font-weight: 800;
      vertical-align: middle;
    }}
    .report-table td {{
      border: 1px solid #cbd5e1;
      padding: 8px 6px;
      text-align: center;
      background: #ffffff;
      vertical-align: middle;
      white-space: nowrap;
      font-variant-numeric: tabular-nums;
    }}
    .report-table tbody tr:nth-child(even) td {{
      background: #f8fafc;
    }}
    .report-table td.fixed {{
      font-weight: 800;
      color: #172033;
    }}
    .report-table td.rank-high {{
      background: #ecfdf3 !important;
      color: #116329;
      font-weight: 800;
    }}
    .report-table td.rank-low {{
      background: #fff1f0 !important;
      color: #b42318;
      font-weight: 800;
    }}
    .charts {{
      display: grid;
      grid-template-columns: repeat(3, 1fr);
      gap: 12px;
    }}
    .chart-section {{
      margin-top: 16px;
    }}
    .section-title {{
      display: flex;
      align-items: baseline;
      justify-content: space-between;
      gap: 16px;
      margin: 0 0 10px;
    }}
    .section-title h2 {{
      margin: 0;
      color: #102a43;
      font-size: 22px;
      line-height: 1.2;
    }}
    .section-title span {{
      color: #52606d;
      font-size: 14px;
    }}
    .chart-card {{
      background: #ffffff;
      border: 1px solid #d9e2ec;
      border-radius: 6px;
      padding: 12px 12px 14px;
    }}
    .chart-card h2 {{
      margin: 0 0 12px;
      color: #102a43;
      font-size: 16px;
      line-height: 1.2;
    }}
    .chart-columns {{
      display: grid;
      grid-template-columns: 1fr 1fr;
      gap: 10px;
    }}
    .chart-label {{
      margin-bottom: 7px;
      font-size: 13px;
      line-height: 1.2;
      font-weight: 900;
    }}
    .chart-label-top {{
      color: #116329;
    }}
    .chart-label-last {{
      color: #b42318;
    }}
    .bar-row {{
      display: grid;
      grid-template-columns: 72px 1fr 54px;
      align-items: center;
      gap: 6px;
      margin: 7px 0;
      font-size: 12px;
    }}
    .bar-label {{
      color: #334e68;
      font-weight: 700;
      white-space: nowrap;
      overflow: hidden;
      text-overflow: ellipsis;
    }}
    .bar-track {{
      height: 12px;
      background: #e2e8f0;
      border-radius: 999px;
      overflow: hidden;
    }}
    .bar-fill {{
      height: 100%;
      border-radius: 999px;
    }}
    .bar-fill-top {{
      background: linear-gradient(90deg, #1d4e89, #38bdf8);
    }}
    .bar-fill-last {{
      background: linear-gradient(90deg, #b42318, #f97316);
    }}
    .bar-value {{
      color: #102a43;
      font-weight: 800;
      text-align: right;
      font-variant-numeric: tabular-nums;
    }}
    .footnote {{
      margin-top: 12px;
      color: #52606d;
      font-size: 13px;
      line-height: 1.6;
    }}
  </style>
</head>
<body>
  <main class="page">
    <header class="title-row">
      <div>
        <h1>{escape(city_name)}地市日通报</h1>
        <div class="subtitle">数据日期：{escape(title_date)}　账期：{escape(month_id or "-")}　区县数：{len(rows)}</div>
      </div>
      <div class="meta">
        <div>生成时间：{escape(generated_at)}</div>
        <div>数据来源：{escape(source.name)}</div>
      </div>
    </header>

    <section class="table-wrap">
      {table}
    </section>

    {focus_charts}

    <div class="footnote">
      上方宽表展示该地市各区县全量业务字段；下方柱状图展示 151装维量、FTTR-H/B装维量、人均价值积分 TOP 5 和 LAST 5。地市日通报只做 HTML 查询，不推送企业微信。
    </div>
  </main>
</body>
</html>
"""


def default_output_path(root: Path, city: str, source: Path) -> Path:
    acct_day, _ = parse_dates(source)
    suffix = acct_day or datetime.now().strftime("%Y%m%d")
    safe_city = re.sub(r'[\\/:*?"<>|\\s]+', "_", normalize_city(city)).strip("_") or "unknown"
    return root / "output" / f"地市日通报_{safe_city}_{suffix}.html"


def main() -> None:
    parser = argparse.ArgumentParser(description="生成地市日通报 HTML；基于各区县装维日清单筛选某地市所有区县")
    parser.add_argument("--city", required=True, help="地市名称，例如 聊城/聊城市")
    parser.add_argument("--input", help="区县汇总 Excel；默认读取 temp/data/city 本轮下载 xlsx")
    parser.add_argument("--city-dir", default="temp/data/city")
    parser.add_argument("--output", help="HTML 输出路径；默认 output/地市日通报_{地市}_{acct_day}.html")
    args = parser.parse_args()

    root = project_root()
    if args.input:
        source = Path(args.input)
        if not source.is_absolute():
            source = root / source
    else:
        source = latest_xlsx(root / args.city_dir)
        if not source:
            raise SystemExit(f"未找到区县汇总 Excel: {root / args.city_dir}")

    headers, raw_rows = load_rows(source, args.city)
    rows = aggregate_rows(headers, raw_rows)
    if not rows:
        raise SystemExit(f"未找到地市={args.city} 的区县汇总数据。")
    rows.sort(key=lambda row: normalize_county(row.get("区县")))

    output = Path(args.output) if args.output else default_output_path(root, args.city, source)
    if not output.is_absolute():
        output = root / output
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(render_html(source, args.city, headers, rows), encoding="utf-8")

    acct_day, month_id = parse_dates(source)
    print("=== 地市日通报 HTML ===")
    print(f"来源文件: {source}")
    print(f"地市: {normalize_city(args.city)}")
    print(f"日期参数: month_id={month_id or '-'}, acct_day={acct_day or '-'}")
    print(f"区县行数: {len(rows)}")
    print(f"输出文件: {output}")


if __name__ == "__main__":
    main()
