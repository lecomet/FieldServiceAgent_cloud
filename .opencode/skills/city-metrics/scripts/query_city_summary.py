#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
import sys
import warnings
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

try:
    import openpyxl
except ImportError as exc:
    raise SystemExit("缺少依赖 openpyxl，请使用 uv --with openpyxl 运行。") from exc

warnings.filterwarnings("ignore", message="Workbook contains no default style.*", category=UserWarning)

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")


REQUIRED_COLUMNS = {"地市", "区县", "装维发展量", "全量发展量", "装维占比"}
RATIO_COLUMNS = {"装维占比": ("装维发展量", "全量发展量")}
PREVIEW_GROUPS = [
    ["地市", "区县", "宽带", "移动", "151"],
    ["地市", "区县", "FTTR-H", "FTTR-B", "FTTR-H/B"],
    ["地市", "区县", "装维发展量", "全量发展量", "装维占比"],
    ["地市", "区县", "发展积分", "运营积分", "原始积分"],
]
DEFAULT_MAX_OUTPUT_ROWS = 20


def project_root() -> Path:
    return Path(__file__).resolve().parents[4]


def latest_xlsx(directory: Path) -> Path | None:
    files = [path for path in directory.glob("*.xlsx") if path.is_file() and not path.name.startswith("~$")]
    if not files:
        return None
    return max(files, key=lambda path: path.stat().st_mtime)


def clean_header(value) -> str:
    return str(value or "").replace("\r", "").replace("\n", "").replace("\t", "").strip()


def format_fields(headers: list[str]) -> str:
    fields = [header for header in headers if header]
    return "、".join(fields) if fields else "无"


def validate_xlsx_file(path: Path) -> None:
    if path.suffix.lower() != ".xlsx":
        raise ValueError(f"{path} 文件格式不符合要求：只支持 .xlsx")
    with path.open("rb") as handle:
        signature = handle.read(4)
    if not signature.startswith(b"PK"):
        raise ValueError(f"{path} 不是标准 .xlsx 文件。请确认 BI 下载格式为 xlsx，不要把 CSV/HTML/XLS 改名为 .xlsx。")


def read_headers(path: Path, iterator) -> list[str]:
    try:
        raw_headers = next(iterator)
    except StopIteration as exc:
        raise ValueError(f"{path} 文件为空：第 1 行必须是字段名") from exc
    headers = [clean_header(value) for value in raw_headers]
    non_empty = [header for header in headers if header]
    duplicates = sorted({header for header in non_empty if non_empty.count(header) > 1})
    if duplicates:
        raise ValueError(f"{path} 第 1 行字段名重复: {', '.join(duplicates)}\n实际字段: {format_fields(headers)}")
    return headers


def require_columns(path: Path, indexes: dict[str, int], required: set[str], headers: list[str]) -> None:
    missing = sorted(required - set(indexes))
    if missing:
        raise ValueError(f"{path} 缺少字段: {', '.join(missing)}\n实际字段: {format_fields(headers)}")


def normalize_city(value) -> str:
    text = str(value or "").strip()
    return text[:-1] if text.endswith("市") else text


def text_match(value, query: str, normalize=None) -> bool:
    value_text = str(value or "").strip()
    query_text = str(query or "").strip()
    if normalize:
        value_text = normalize(value_text)
        query_text = normalize(query_text)
    return bool(value_text and query_text and (value_text == query_text or query_text in value_text or value_text in query_text))


def number(value) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    text = str(value).strip().replace(",", "")
    if not text:
        return Decimal("0")
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal("0")


def load_rows(path: Path, city: str | None, county: str | None) -> tuple[list[str], list[dict]]:
    validate_xlsx_file(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    iterator = ws.iter_rows(values_only=True)
    try:
        headers = read_headers(path, iterator)
        indexes = {header: idx for idx, header in enumerate(headers) if header}
        require_columns(path, indexes, REQUIRED_COLUMNS, headers)

        rows = []
        for raw in iterator:
            if city and not text_match(raw[indexes["地市"]], city, normalize_city):
                continue
            if county and not text_match(raw[indexes["区县"]], county):
                continue
            row = {header: raw[idx] if idx < len(raw) else None for header, idx in indexes.items()}
            row["地市"] = normalize_city(row.get("地市"))
            rows.append(row)
    finally:
        wb.close()
    return headers, rows


def aggregate_rows(headers: list[str], rows: list[dict]) -> list[dict]:
    grouped: dict[tuple[str, str], dict] = {}
    for row in rows:
        city = normalize_city(row.get("地市"))
        county = str(row.get("区县") or "").strip()
        if not city or not county:
            continue
        target = grouped.setdefault((city, county), {header: None for header in headers})
        target["地市"] = city
        target["区县"] = county
        for header in headers:
            if header in {"地市", "区县"} or header in RATIO_COLUMNS:
                continue
            target[header] = number(target.get(header)) + number(row.get(header))

    for row in grouped.values():
        for ratio_col, (num_col, den_col) in RATIO_COLUMNS.items():
            if ratio_col not in headers:
                continue
            denominator = number(row.get(den_col))
            row[ratio_col] = number(row.get(num_col)) / denominator if denominator else Decimal("0")
    return list(grouped.values())


def format_value(value) -> str:
    if isinstance(value, Decimal):
        return str(round(float(value), 4)).rstrip("0").rstrip(".")
    if isinstance(value, float):
        return str(round(value, 4)).rstrip("0").rstrip(".")
    return str(value or "")


def markdown_table(headers: list[str], rows: list[dict]) -> str:
    lines = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join("---" for _ in headers) + " |",
    ]
    for row in rows:
        lines.append("| " + " | ".join(format_value(row.get(header)) for header in headers) + " |")
    return "\n".join(lines)


def scope_token(city: str | None, county: str | None) -> str:
    if county:
        return f"county_{county}"
    if city:
        return f"city_{normalize_city(city)}"
    return "all_county"


def write_excel(output_path: Path, headers: list[str], rows: list[dict], source_file: Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "区县汇总"
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header) for header in headers])

    meta = wb.create_sheet("来源")
    meta.append(["项目", "值"])
    meta.append(["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    meta.append(["来源文件", str(source_file)])
    meta.append(["地市名称处理", "去掉末尾“市”；同一地市同一区县重复行合并；占比按合并后的分子/分母重算"])
    meta.append(["行数", len(rows)])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(output_path)
        return output_path
    except PermissionError:
        fallback = output_path.with_name(f"{output_path.stem}_{datetime.now().strftime('%Y%m%d%H%M%S')}{output_path.suffix}")
        wb.save(fallback)
        return fallback


def main() -> None:
    parser = argparse.ArgumentParser(description="装维区县汇总查询和地市名称标准化")
    parser.add_argument("--city", help="可选，限定地市；会自动兼容泰安/泰安市")
    parser.add_argument("--county", help="可选，限定区县")
    parser.add_argument("--city-dir", default="temp/data/city")
    parser.add_argument("--output", help="可选，输出 Excel 路径")
    parser.add_argument("--max-output", type=int, default=DEFAULT_MAX_OUTPUT_ROWS)
    args = parser.parse_args()

    root = project_root()
    source_file = latest_xlsx(root / args.city_dir)
    if not source_file:
        raise SystemExit(f"未找到区县汇总 Excel: {root / args.city_dir}")

    headers, raw_rows = load_rows(source_file, args.city, args.county)
    rows = aggregate_rows(headers, raw_rows)
    if not rows:
        raise SystemExit("未找到匹配的区县汇总数据。")

    rows.sort(key=lambda row: (normalize_city(row.get("地市")), str(row.get("区县") or "")))
    saved = None
    if args.output:
        output = Path(args.output)
        if not output.is_absolute():
            output = root / output
        saved = write_excel(output, headers, rows, source_file)

    print("=== 装维区县汇总 ===")
    print(f"来源文件: {source_file}")
    if saved:
        print(f"输出文件: {saved}")
    else:
        print(f"输出文件: 未另存；区县汇总下载文件保留在 {args.city_dir}/。")
    print(f"区县行数: {len(rows)}")
    print("地市名称处理: 去掉末尾“市”；同一地市同一区县重复行合并；占比按合并后的分子/分母重算")

    if len(rows) > args.max_output:
        if saved:
            print(f"行数超过 {args.max_output}，不在消息栏罗列；请打开 Excel 查看完整结果。")
        else:
            print(f"行数超过 {args.max_output}，不在消息栏罗列；如需保存格式化副本，请显式传 --output。")
        return

    for group in PREVIEW_GROUPS:
        available = [header for header in group if header in headers]
        if len(available) >= 2:
            print()
            print(markdown_table(available, rows))


if __name__ == "__main__":
    main()
